from flask import Blueprint, jsonify, request, send_file
from sqlalchemy import text
from io import BytesIO
from backend.utils.auth_middleware import require_role, ROL_ADMINS
from backend.models.sql_models import db, ProduccionPulido, PncInyeccion, PncPulido, PncEnsamble, BujeRevuelto, Producto, TrazabilidadLote
from backend.utils.formatters import normalizar_codigo, preservar_o_normalizar_prefijo
import uuid
from datetime import datetime
import pytz
import logging
import json

logger = logging.getLogger(__name__)
pulido_bp = Blueprint('pulido_bp', __name__)

@pulido_bp.route('/api/pulido', methods=['POST'])
def registrar_pulido():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "No data provided"}), 400

        colombia_tz = pytz.timezone('America/Bogota')
        ahora = datetime.now(colombia_tz)

        # Lógica de Upsert (Evitar duplicados por id_pulido)
        id_pulido = data.get('id_pulido')
        registro = ProduccionPulido.query.filter_by(id_pulido=id_pulido).first() if id_pulido else None

        # Guard de ownership: evita que un operario sobrescriba el registro de otro
        from backend.utils.formatters import resolver_operario
        incoming_responsable = resolver_operario(data.get('responsable'))
        if registro and registro.responsable and incoming_responsable:
            if str(registro.responsable).strip().upper() != incoming_responsable.upper():
                return jsonify({
                    "success": False,
                    "error": "La sesión pertenece a otro operario. Cierra sesión y vuelve a ingresar.",
                    "code": "PULIDO_SESSION_OWNERSHIP_MISMATCH",
                    "id_pulido": id_pulido,
                    "responsable_db": registro.responsable,
                    "responsable_in": incoming_responsable
                }), 409

        if not registro:
            # Si no existe (o fue borrado de la DB), crear uno nuevo para evitar Error 500
            if id_pulido:
                logger.warning(f" [RECOVERY] id_pulido {id_pulido} no encontrado en DB. Creando nuevo registro.")
            registro = ProduccionPulido(id_pulido=id_pulido or f"PUL-{ahora.strftime('%Y%m%d%H%M%S')}")
            db.session.add(registro)

        # Mapeo y Estandarización
        registro.fecha = datetime.strptime(data.get('fecha_inicio', ahora.strftime('%Y-%m-%d')), '%Y-%m-%d').date()
        # ── Blindaje Dual: preservar prefijo MT-/CAR- o agregar FR- por defecto ──
        registro.codigo = preservar_o_normalizar_prefijo(data.get('codigo_producto'))
        registro.responsable = incoming_responsable
        registro.cantidad_real = float(data.get('cantidad_real') or 0)
        registro.pnc_inyeccion = int(data.get('pnc_inyeccion') or 0)
        registro.pnc_pulido = int(data.get('pnc_pulido') or 0)
        registro.criterio_pnc_inyeccion = data.get('criterio_pnc_inyeccion')
        registro.criterio_pnc_pulido = data.get('criterio_pnc_pulido')
        registro.orden_produccion = data.get('orden_produccion') or 'SIN OP'
        registro.observaciones = data.get('observaciones', '')
        registro.estado = data.get('estado', 'FINALIZADO')
        registro.departamento = 'Pulido'  # Estandarización exigida
        registro.lote = data.get('lote') or 'SIN LOTE'
        registro.cantidad_recibida = float(data.get('cantidad_recibida') or 0)
        registro.almacen_destino = data.get('almacen_destino', 'P. TERMINADO')

        # Validación de consistencia (Bujes Buenos + PNC <= Total)
        # Se usa <= porque puede haber otros tipos de PNC no mapeados a columnas fijas
        total_reportado = registro.cantidad_real + registro.pnc_inyeccion + registro.pnc_pulido
        if registro.cantidad_recibida < total_reportado:
            logger.warning(f" [VALIDATION] Inconsistencia en {id_pulido}: Total {registro.cantidad_recibida} < Suma {total_reportado}")
            # No bloqueamos el guardado para evitar pérdida de datos en planta, pero registramos el evento.
            # Opcional: podrías retornar error 400 aquí si prefieres rigidez total.


        # Manejo de Horas y Cálculos de Tiempo
        colombia_tz = pytz.timezone('America/Bogota')
        ahora = datetime.now(colombia_tz)

        if data.get('hora_inicio'):
            h_h, h_m = data['hora_inicio'].split(':')
            dt_ini = ahora.replace(hour=int(h_h), minute=int(h_m), second=0, microsecond=0)
            registro.hora_inicio = dt_ini.replace(tzinfo=None) # Guardar como naive Bogota
        
        if data.get('hora_fin'):
            h_h, h_m = data['hora_fin'].split(':')
            dt_fin = ahora.replace(hour=int(h_h), minute=int(h_m), second=0, microsecond=0)
            registro.hora_fin = dt_fin.replace(tzinfo=None) # Guardar como naive Bogota

        # Cálculo de Métricas (Usando objetos localizados para precisión)
        if data.get('hora_inicio') and data.get('hora_fin'):
            hi_h, hi_m = data['hora_inicio'].split(':')
            hf_h, hf_m = data['hora_fin'].split(':')
            
            t_ini = ahora.replace(hour=int(hi_h), minute=int(hi_m), second=0, microsecond=0)
            t_fin = ahora.replace(hour=int(hf_h), minute=int(hf_m), second=0, microsecond=0)
            
            # Tiempo del segmento actual
            diff = t_fin - t_ini
            segundos_segmento = int(diff.total_seconds())
            if segundos_segmento < 0: segundos_segmento += 86400
            
            # Sumar tiempo acumulado de sesiones previas (enviado desde el frontend en ms)
            tiempo_acumulado_ms = float(data.get('tiempo_acumulado_ms') or 0)
            segundos_totales = segundos_segmento + int(tiempo_acumulado_ms / 1000)

            # Descuento automático por pausas programadas (enviado por frontend, en ms)
            descuento_programado_ms = float(data.get('descuento_programado_ms') or 0)
            if descuento_programado_ms > 0:
                segundos_totales = max(0, segundos_totales - int(descuento_programado_ms / 1000))
            
            registro.duracion_segundos = segundos_totales
            registro.tiempo_total_minutos = float(round(segundos_totales / 60.0, 2))
            
            cant = float(registro.cantidad_real or 0)
            if cant > 0:
                registro.segundos_por_unidad = float(round(segundos_totales / cant, 2))
            else:
                registro.segundos_por_unidad = 0.0
            
            # Persistir evidencia del descuento en observaciones (sin migración de DB)
            try:
                detalle = data.get('detalle_descuento_programado')
                if isinstance(detalle, str):
                    detalle = json.loads(detalle)
                if isinstance(detalle, list) and descuento_programado_ms > 0:
                    payload = {
                        "descuento_programado_min": round(descuento_programado_ms / 60000.0, 2),
                        "detalle": detalle
                    }
                    tag = f"[AUTO_BREAK]{json.dumps(payload, ensure_ascii=False)}[/AUTO_BREAK]"
                    obs = (registro.observaciones or "")
                    # Reemplazar tag si ya existía
                    if "[AUTO_BREAK]" in obs and "[/AUTO_BREAK]" in obs:
                        pre = obs.split("[AUTO_BREAK]")[0]
                        post = obs.split("[/AUTO_BREAK]")[-1]
                        registro.observaciones = (pre + tag + post).strip()
                    else:
                        registro.observaciones = (obs + "\n" + tag).strip() if obs else tag
            except Exception as e:
                logger.warning(f"[AUTO_BREAK] No se pudo persistir detalle: {e}")

            logger.info(f" [TIME-DEBUG] {id_pulido} -> Seg: {segundos_segmento}s, Acum: {tiempo_acumulado_ms}ms, DescProg: {descuento_programado_ms}ms, Total: {segundos_totales}s")
        else:
            registro.duracion_segundos = 0
            registro.tiempo_total_minutos = 0.0
            registro.segundos_por_unidad = 0.0

        db.session.flush() # Asegurar que el registro principal tenga ID en la sesión
 
        # Sincronización de PNC Detallado (vaya esa info donde debe)
        pnc_detail = data.get('pnc_detail', [])
        if pnc_detail:
            # Limpiar registros previos de PNC vinculados a este id_pulido
            # Usar registro.id_pulido que es el identificador confirmado en este punto
            db.session.query(PncInyeccion).filter_by(id_inyeccion=registro.id_pulido).delete()
            db.session.query(PncPulido).filter_by(id_pulido=registro.id_pulido).delete()
            db.session.query(PncEnsamble).filter_by(id_ensamble=registro.id_pulido).delete()

            for pnc_item in pnc_detail:
                proc = pnc_item.get('proceso', '').upper()
                cant = float(pnc_item.get('cantidad') or 0)
                crit = pnc_item.get('criterio', '')
                if cant <= 0: continue

                if proc == 'INYECCION':
                    db.session.add(PncInyeccion(
                        id_pnc_inyeccion=uuid.uuid4().hex[:8],
                        id_inyeccion=registro.id_pulido,
                        id_codigo=registro.codigo,
                        cantidad=cant,
                        criterio=crit
                    ))
                elif proc == 'PULIDO':
                    db.session.add(PncPulido(
                        id_pnc_pulido=uuid.uuid4().hex[:8],
                        id_pulido=registro.id_pulido,
                        codigo=registro.codigo,
                        cantidad=cant,
                        criterio=crit
                    ))
                elif proc == 'ENSAMBLE':
                    db.session.add(PncEnsamble(
                        id_pnc_ensamble=uuid.uuid4().hex[:8],
                        id_ensamble=registro.id_pulido,
                        id_codigo=registro.codigo,
                        cantidad=cant,
                        criterio=crit
                    ))

        db.session.flush() # Asegurar que el registro principal tenga ID en la sesión

        # ---------------------------------------------------------
        # Manejo de Bujes Revueltos (NUEVO)
        # ---------------------------------------------------------
        revueltos_list = data.get('revueltos', [])
        logger.info(f" [REVUELTOS-DEBUG] Payload recibido: {revueltos_list}")
        logger.info(f" [REVUELTOS] Procesando {len(revueltos_list)} items para {registro.id_pulido}")
        
        # Limpiar registros previos de revueltos vinculados a este id_pulido
        db.session.query(BujeRevuelto).filter_by(id_pulido=registro.id_pulido).delete()

        for rev_item in revueltos_list:
            cod_rev = preservar_o_normalizar_prefijo(rev_item.get('id_codigo'))  # Prefijo obligatorio
            cant_rev = float(rev_item.get('cantidad') or 0)
            logger.info(f" [REVUELTOS] Intentando agregar: {cod_rev} | Cant: {cant_rev}")
            
            if cant_rev <= 0 or not cod_rev: 
                logger.warning(f" [REVUELTOS] Item omitido por validación: {rev_item}")
                continue

            db.session.add(BujeRevuelto(
                id_bujes_revueltos=uuid.uuid4().hex[:8],
                id_pulido=registro.id_pulido,
                id_codigo=cod_rev,
                cantidad=cant_rev,
                codigo_ensamble=cod_rev,
                responsable=registro.responsable
            ))
            logger.info(f" [REVUELTOS] Registro agregado a sesión: {cod_rev}")

        # --- Propagación de avances a cubetas FIFO (db_distribucion_op_pedidos) ---
        op_actual = registro.orden_produccion
        if op_actual and str(op_actual).strip() != 'SIN OP':
            from backend.models.sql_models import DistribucionOpPedidos
            
            # ── Blindaje Dual ──
            # op_limpia   → tal cual (db_distribucion_op_pedidos no usa prefijo)
            # codigo_limpio → SIN prefijo (db_distribucion_op_pedidos guarda '9890')
            # registro.codigo ya tiene FR- (escrito arriba con con_prefijo_fr)
            op_limpia = str(registro.orden_produccion or '').strip()
            codigo_limpio = normalizar_codigo(registro.codigo)  # quita FR- para el cruce
            
            # Buscar las cubetas por OP y Referencia ordenadas de forma ascendente
            cubetas = db.session.query(DistribucionOpPedidos).filter(
                DistribucionOpPedidos.op_world_office == op_limpia,
                DistribucionOpPedidos.codigo_producto == codigo_limpio
            ).order_by(DistribucionOpPedidos.id_distribucion.asc()).all()

            piezas_por_repartir = float(registro.cantidad_real or 0)
            
            # Validación y creación de cubeta de contingencia
            if not cubetas and piezas_por_repartir > 0:
                # Intentar buscar el id_pedido de otra cubeta asociada a la misma OP
                pedido_asoc = db.session.query(DistribucionOpPedidos.id_pedido).filter(
                    DistribucionOpPedidos.op_world_office == op_limpia
                ).first()
                id_pedido_final = pedido_asoc[0] if (pedido_asoc and pedido_asoc[0]) else f"PED-IMPREVISTO-{op_limpia}"
                
                logger.info(f" ⚠️ [PULIDO-CONTINGENCIA] Creando cubeta temporal para OP: {op_limpia}, Producto: {codigo_limpio}, Pedido: {id_pedido_final}")
                nueva_cubeta = DistribucionOpPedidos(
                    op_world_office=op_limpia,
                    id_pedido=id_pedido_final,
                    codigo_producto=codigo_limpio,
                    cant_requerida=piezas_por_repartir,
                    cant_inyectada=piezas_por_repartir, # Nivelar inyección
                    cant_pulida=piezas_por_repartir,
                    cant_ensamblada=0,
                    cant_alistada=0
                )
                db.session.add(nueva_cubeta)
                db.session.flush() # Sincronizar temporalmente en sesión
                cubetas = [nueva_cubeta]
                piezas_por_repartir = 0.0 # Consumido por completo
            
            logger.info(f" 📦 [PULIDO-FIFO] Propagando {piezas_por_repartir} piezas a {len(cubetas)} cubetas. OP: {op_limpia}, Producto: {codigo_limpio}")
            
            for cubeta in cubetas:
                if piezas_por_repartir <= 0:
                    break
                
                # Cuánto le falta a esta cubeta en la etapa de pulido
                falta = max(0, (cubeta.cant_requerida or 0) - (cubeta.cant_pulida or 0))
                if falta > 0:
                    if piezas_por_repartir >= falta:
                        cubeta.cant_pulida = (cubeta.cant_pulida or 0) + falta
                        piezas_por_repartir -= falta
                    else:
                        cubeta.cant_pulida = (cubeta.cant_pulida or 0) + piezas_por_repartir
                        piezas_por_repartir = 0

        # ---------------------------------------------------------
        # Cierre de Lote con Saldo a WIP "Por Pulir"
        # ---------------------------------------------------------
        lote_traz = db.session.get(TrazabilidadLote, registro.id_pulido)
        if lote_traz:
            lote_traz.estado_actual = 'PENDIENTE_VALIDACION'
            
            cant_inyectada = float(lote_traz.cantidad_inyectada or 0)
            buenas = float(registro.cantidad_real or 0)
            pnc_total = float(registro.pnc_inyeccion or 0) + float(registro.pnc_pulido or 0)
            rev_total = sum(float(r.get('cantidad', 0)) for r in revueltos_list)
            
            wip = cant_inyectada - (buenas + pnc_total + rev_total)
            
            if wip > 0:
                es_prueba_wip = '9999' in str(registro.orden_produccion).upper() or 'PRUEBA' in str(registro.orden_produccion).upper() or '9999' in str(registro.id_pulido).upper() or 'PRUEBA' in str(registro.id_pulido).upper() or '9999' in str(lote_traz.id_lote).upper() or 'PRUEBA' in str(lote_traz.id_lote).upper()
                
                if not es_prueba_wip:
                    # ── WIP: buscar en db_productos respetando prefijo ──
                    prod_wip = db.session.query(Producto).filter_by(
                        codigo_sistema=preservar_o_normalizar_prefijo(registro.codigo)
                    ).first()
                    if prod_wip:
                        prod_wip.por_pulir = float(prod_wip.por_pulir or 0) + wip
                        logger.info(f" [WIP] Agregando {wip} piezas a por_pulir del producto {registro.codigo}")
                else:
                    logger.info(f"🧪 [SANDBOX] Lote de prueba {registro.id_pulido}. Se ignoró suma de {wip} al WIP (por_pulir).")

        db.session.commit()
        return jsonify({
            "success": True, 
            "message": "Registro de pulido sincronizado",
            "id_pulido": registro.id_pulido,
            "upsert": "UPDATE" if id_pulido and registro.id else "INSERT"
        }), 201

    except Exception as e:
        db.session.rollback()
        import traceback
        error_trace = traceback.format_exc()
        logger.error(f"❌ [PULIDO-ERROR] Error crítico al registrar: {str(e)}\n{error_trace}")
        return jsonify({
            "success": False, 
            "error": str(e),
            "detail": "Error interno al procesar el reporte de pulido (Revisa logs del servidor)"
        }), 500

@pulido_bp.route('/api/debug/forzar_a_pulido/<id_lote>', methods=['GET', 'POST'])
def debug_forzar_pulido(id_lote):
    try:
        from backend.models.sql_models import TrazabilidadLote
        lote = TrazabilidadLote.query.filter_by(id_lote=id_lote).first()
        if not lote:
            return jsonify({"success": False, "error": f"Lote {id_lote} no encontrado"}), 404
        
        lote.estado_actual = 'ABIERTO_PULIDO'
        if not lote.por_pulir or lote.por_pulir <= 0:
            lote.por_pulir = 100
        
        db.session.commit()
        logger.info(f"🛠️ [DEBUG] Lote {id_lote} forzado a ABIERTO_PULIDO con por_pulir={lote.por_pulir}")
        return jsonify({"success": True, "message": f"Lote {id_lote} forzado a ABIERTO_PULIDO"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

@pulido_bp.route('/api/pulido/session_active', methods=['GET'])
def get_active_pulido_session():
    try:
        # Forzar limpieza de caché de sesión para obtener datos frescos de la DB
        db.session.remove()
        
        if request.args.get('ping') == 'true':
            return jsonify({"success": True, "ping": "pong"}), 200

        responsable = request.args.get('responsable')
        if not responsable:
            return jsonify({"success": False, "error": "Falta responsable"}), 400
        
        # FIX Efecto Daniela: Filtrar por estado EN_PROCESO/PAUSADO,
        # NO por hora_fin=None (registros migrados no tienen hora_fin)
        query = ProduccionPulido.query.filter(ProduccionPulido.responsable == responsable)
        
        id_especifico = request.args.get('id_pulido')
        if id_especifico:
            query = query.filter(ProduccionPulido.id_pulido == id_especifico)
        else:
            query = query.filter(ProduccionPulido.estado.in_(['EN_PROCESO', 'PAUSADO', 'PAUSADO_COLA', 'TRABAJANDO']))
            
        sesion = query.order_by(ProduccionPulido.id.desc()).first()

        if sesion:
            return jsonify({
                "success": True,
                "session": {
                    "id_pulido": sesion.id_pulido,
                    "codigo": sesion.codigo,
                    "lote": sesion.lote,
                    "orden_produccion": sesion.orden_produccion,
                    "hora_inicio_dt": sesion.hora_inicio.isoformat() if sesion.hora_inicio else None,
                    "duracion_segundos": sesion.duracion_segundos or 0,
                    "estado": sesion.estado,
                    "tiempo_pausa_acumulado": sesion.tiempo_pausa_acumulado or 0,
                    "hora_pausa": sesion.hora_pausa.isoformat() if (sesion.estado == 'PAUSADO' and sesion.hora_pausa) else None
                }
            })
        
        return jsonify({"success": True, "session": None})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@pulido_bp.route('/api/pulido/tareas_pendientes', methods=['GET'])
def get_pulido_tareas_pendientes():
    try:
        responsable = request.args.get('responsable')
        if not responsable:
            return jsonify({"success": False, "error": "Falta responsable"}), 400
        
        tareas = ProduccionPulido.query.filter(
            ProduccionPulido.responsable == responsable,
            ProduccionPulido.estado.in_(['PENDIENTE', 'PAUSADO_COLA'])
        ).order_by(ProduccionPulido.id.desc()).all()
        
        return jsonify({
            "success": True,
            "tareas": [{
                "id_pulido": t.id_pulido,
                "codigo": t.codigo,
                "lote": t.lote,
                "orden_produccion": t.orden_produccion,
                "estado": t.estado
            } for t in tareas]
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@pulido_bp.route('/api/pulido/pausar', methods=['POST'])
def pausar_pulido():
    """Registra el inicio de la pausa en el servidor."""
    data = request.json
    id_pulido = data.get('id_pulido')
    try:
        registro = ProduccionPulido.query.filter_by(id_pulido=id_pulido).first()
        if not registro: return jsonify({"success": False, "error": "No encontrado"}), 404
        
        # Blindaje: Forzar timestamp de Colombia (Bogotá)
        colombia_tz = pytz.timezone('America/Bogota')
        ahora = datetime.now(colombia_tz)

        # Si el frontend envía una hora específica, intentar usarla para la parte de tiempo
        hora_front = data.get('hora_pausa')
        if hora_front and ':' in hora_front:
            try:
                h, m = hora_front.split(':')
                ahora = ahora.replace(hour=int(h), minute=int(m), second=0, microsecond=0)
            except: pass

        registro.estado = 'PAUSADO'
        registro.hora_pausa = ahora.replace(tzinfo=None) # Guardar como naive Bogota
        db.session.add(registro)
        db.session.commit()
        
        logger.info(f" [PAUSA] Actividad {id_pulido} pausada a las {registro.hora_pausa}")
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

@pulido_bp.route('/api/pulido/reanudar', methods=['POST'])
def reanudar_pulido():
    """Calcula el tiempo de la pausa y lo suma al acumulador."""
    data = request.json
    id_pulido = data.get('id_pulido')
    try:
        registro = ProduccionPulido.query.filter_by(id_pulido=id_pulido).first()
        if not registro: return jsonify({"success": False, "error": "No encontrado"}), 404
        
        if registro.hora_pausa:
            colombia_tz = pytz.timezone('America/Bogota')
            ahora = datetime.now(colombia_tz)

            # Si el frontend envía una hora específica de reanudación
            hora_front = data.get('hora_reanudar')
            if hora_front and ':' in hora_front:
                try:
                    h, m = hora_front.split(':')
                    ahora = ahora.replace(hour=int(h), minute=int(m), second=0, microsecond=0)
                except: pass

            ahora_naive = ahora.replace(tzinfo=None)
            diferencia = ahora_naive - registro.hora_pausa
            segundos_pausa = int(diferencia.total_seconds())
            if segundos_pausa < 0: segundos_pausa = 0 # Evitar pausas negativas por drift
            
            registro.tiempo_pausa_acumulado = (registro.tiempo_pausa_acumulado or 0) + segundos_pausa
            
        registro.estado = 'TRABAJANDO'
        registro.hora_pausa = None
        db.session.commit()
        return jsonify({"success": True, "acumulado": registro.tiempo_pausa_acumulado})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

@pulido_bp.route('/api/pulido/swap_task', methods=['POST'])
def swap_pulido_task():
    try:
        data = request.get_json()
        responsable = str(data.get('responsable') or "")
        id_raw = data.get('id_pulido')

        # Registro de depuración para Render
        logger.info(f" [SWAP-DEBUG] Recibido id_pulido: {id_raw} (Tipo: {type(id_raw)})")

        # Blindaje definitivo contra [object Object] o diccionarios
        if isinstance(id_raw, dict):
            id_nuevo = str(id_raw.get('id_pulido') or "")
        elif str(id_raw) == "[object Object]":
            logger.error(" [SWAP-ERROR] El frontend envió un string '[object Object]'")
            return jsonify({"success": False, "error": "ID de tarea inválido (object)"}), 400
        else:
            id_nuevo = str(id_raw or "")
        
        if not responsable or not id_nuevo or id_nuevo == "None":
            return jsonify({"success": False, "error": "Datos incompletos o ID nulo"}), 400

        # 1. Pausar TODO lo que esté TRABAJANDO para este operario
        # Usamos la hora de Colombia para asegurar consistencia en el registro de pausa
        colombia_tz = pytz.timezone('America/Bogota')
        now = datetime.now(colombia_tz).replace(tzinfo=None)
        trabajos_activos = ProduccionPulido.query.filter(
            ProduccionPulido.responsable == responsable,
            ProduccionPulido.estado == 'TRABAJANDO'
        ).all()
        
        for t in trabajos_activos:
            t.estado = 'PAUSADO_COLA'
            t.hora_pausa = now
            db.session.add(t)
            
        # 2. Activar la nueva tarea
        nueva_tarea = ProduccionPulido.query.filter_by(id_pulido=id_nuevo).first()
        if nueva_tarea:
            nueva_tarea.estado = 'TRABAJANDO'
            nueva_tarea.hora_pausa = None # Limpiar pausa para reanudación
            db.session.add(nueva_tarea)
            
        db.session.commit()
        return jsonify({"success": True, "message": "Intercambio realizado con éxito"})
    except Exception as e:
        db.session.rollback()
        logger.error(f"[Pulido-Swap] Error crítico: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
@pulido_bp.route('/api/pulido/historial', methods=['GET'])
def get_pulido_historial():
    """
    Endpoint para el historial detallado de Pulido con filtros dinámicos.
    """
    try:
        # 1. Parámetros
        f_inicio = request.args.get('fecha_inicio', '')
        f_fin = request.args.get('fecha_fin', '')
        operario = request.args.get('operario', '')
        id_codigo = request.args.get('id_codigo', '')

        # 2. Base Query con Casts explícitos para evitar OID 25 (TEXT)
        sql = """
            SELECT 
                id,
                id_pulido::TEXT as id_pulido,
                fecha,
                codigo::TEXT as codigo,
                responsable::TEXT as responsable,
                cantidad_real,
                pnc_inyeccion,
                pnc_pulido,
                hora_inicio,
                hora_fin,
                orden_produccion::TEXT as orden_produccion,
                observaciones::TEXT as observaciones,
                cantidad_recibida
            FROM db_pulido
            WHERE 1=1
        """
        params = {}

        # 3. Filtros Dinámicos
        if f_inicio and f_fin:
            sql += " AND CAST(fecha AS DATE) BETWEEN :f_inicio AND :f_fin"
            params['f_inicio'] = f_inicio
            params['f_fin'] = f_fin
        elif f_inicio:
            sql += " AND CAST(fecha AS DATE) >= :f_inicio"
            params['f_inicio'] = f_inicio
        elif f_fin:
            sql += " AND CAST(fecha AS DATE) <= :f_fin"
            params['f_fin'] = f_fin

        if operario and operario.upper() != 'TODOS':
            sql += " AND UPPER(TRIM(responsable)) LIKE :operario"
            params['operario'] = f"%{operario.strip().upper()}%"

        if id_codigo and id_codigo.upper() != 'TODOS':
            sql += " AND UPPER(TRIM(codigo)) LIKE :codigo"
            params['codigo'] = f"%{id_codigo.strip().upper()}%"

        sql += " ORDER BY fecha DESC, id DESC"

        # 4. Ejecución y Conversión Inmediata a Diccionario (Clean Data)
        result = db.session.execute(text(sql), params)
        resultados = [dict(row._mapping) for row in result]

        # 5. Batch Pre-fetch de Revueltos (v5.2 - Zero queries in loop)
        all_ids = [str(r.get('id_pulido') or '').strip() for r in resultados]
        all_ids = [pid for pid in all_ids if pid]  # Filtrar vacíos

        revueltos_map = {}  # { id_pulido: [ {id_codigo, cantidad}, ... ] }
        if all_ids:
            # Una sola consulta para TODOS los revueltos
            placeholders = ', '.join([f':pid_{i}' for i in range(len(all_ids))])
            sql_revs = f"SELECT id_pulido::TEXT as id_pulido, id_codigo::TEXT as id_codigo, COALESCE(cantidad, 0) as cantidad FROM db_bujes_revueltos WHERE id_pulido IN ({placeholders})"
            params_revs = {f'pid_{i}': pid for i, pid in enumerate(all_ids)}
            revs_raw = db.session.execute(text(sql_revs), params_revs)
            for rv in revs_raw:
                rv_dict = dict(rv._mapping)
                pid = str(rv_dict['id_pulido'])
                if pid not in revueltos_map:
                    revueltos_map[pid] = []
                revueltos_map[pid].append(rv_dict)

        # 6. Mapeo para el Frontend (cero consultas en el bucle)
        data = []
        for r in resultados:
            p_id = str(r.get('id_pulido') or '').strip()

            # Lookup directo en memoria
            revs = revueltos_map.get(p_id, [])
            det_revueltos = ""
            if revs:
                det_revueltos = " | REVUELTOS: " + ", ".join([f"{str(rv['id_codigo'])}({float(rv['cantidad'] or 0)})" for rv in revs])

            obs = str(r.get('observaciones') or '').strip()
            detalle_final = f"Obs: {obs}{det_revueltos}" if obs else det_revueltos.strip(" | ")

            data.append({
                'id': int(r['id']),
                'Fecha': r['fecha'].strftime('%d/%m/%Y') if r['fecha'] else '',
                'Tipo': 'PULIDO',
                'Producto': str(r['codigo'] or ''),
                'Responsable': str(r['responsable'] or 'SISTEMA'),
                'cantidad_real': float(r['cantidad_real'] or 0),
                'Cant': float(r['cantidad_real'] or 0),
                'Orden': str(r['orden_produccion'] or p_id or '-'),
                'Extra': str(f"OP: {r['orden_produccion'] or ''}"),
                'Detalle': str(detalle_final.strip()),
                'HORA_INICIO': r['hora_inicio'].strftime('%H:%M') if r['hora_inicio'] else '',
                'HORA_FIN': r['hora_fin'].strftime('%H:%M') if r['hora_fin'] else '',
                'hoja': 'db_pulido',
                'fila': int(r['id']),
                'cantidad_recibida': float(r['cantidad_recibida'] or 0),
                'pnc_inyeccion': int(r['pnc_inyeccion'] or 0),
                'pnc_pulido': int(r['pnc_pulido'] or 0)
            })

        return jsonify(data)

    except Exception as e:
        db.session.rollback()
        logger.error(f"Error en api/pulido/historial: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@pulido_bp.route('/api/pulido/stats', methods=['GET'])
@require_role(ROL_ADMINS + ['JEFE PULIDO', 'PULIDO'])
def get_pulido_stats():
    # Placeholder
    return jsonify({"success": True, "message": "Estadísticas de pulido (WIP)"})


@pulido_bp.route('/api/pulido/exportar_excel', methods=['GET'])
def exportar_excel_pulido():
    """
    Exportación profesional a Excel del historial de Pulido.
    Usa openpyxl para estilos avanzados (cabecera, cebreado, alertas PNC).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    try:
        # 1. Parámetros (mismos filtros que el historial)
        f_inicio = request.args.get('fecha_inicio', '')
        f_fin = request.args.get('fecha_fin', '')
        operario = request.args.get('operario', '')
        id_codigo = request.args.get('id_codigo', '')

        # 2. Query con casts
        sql = """
            SELECT
                id, id_pulido::TEXT as id_pulido, fecha,
                codigo::TEXT as codigo, responsable::TEXT as responsable,
                cantidad_real, pnc_inyeccion, pnc_pulido,
                hora_inicio, hora_fin,
                orden_produccion::TEXT as orden_produccion,
                observaciones::TEXT as observaciones,
                cantidad_recibida
            FROM db_pulido
            WHERE 1=1
        """
        params = {}

        if f_inicio and f_fin:
            sql += " AND CAST(fecha AS DATE) BETWEEN :f_inicio AND :f_fin"
            params['f_inicio'] = f_inicio
            params['f_fin'] = f_fin
        elif f_inicio:
            sql += " AND CAST(fecha AS DATE) >= :f_inicio"
            params['f_inicio'] = f_inicio
        elif f_fin:
            sql += " AND CAST(fecha AS DATE) <= :f_fin"
            params['f_fin'] = f_fin

        if operario and operario.upper() != 'TODOS':
            sql += " AND UPPER(TRIM(responsable)) LIKE :operario"
            params['operario'] = f"%{operario.strip().upper()}%"

        if id_codigo and id_codigo.upper() != 'TODOS':
            sql += " AND UPPER(TRIM(codigo)) LIKE :codigo"
            params['codigo'] = f"%{id_codigo.strip().upper()}%"

        sql += " ORDER BY fecha DESC, id DESC"

        result = db.session.execute(text(sql), params)
        resultados = [dict(row._mapping) for row in result]

        # 3. Crear Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial Pulido"

        # --- ESTILOS ---
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D5D8DC'),
            right=Side(style='thin', color='D5D8DC'),
            top=Side(style='thin', color='D5D8DC'),
            bottom=Side(style='thin', color='D5D8DC')
        )
        zebra_fill = PatternFill(start_color='F2F3F4', end_color='F2F3F4', fill_type='solid')
        pnc_fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
        data_align = Alignment(horizontal='center', vertical='center')
        text_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # --- CABECERA ---
        columnas = [
            'Fecha', 'Hora Inicio', 'Hora Fin', 'Operario', 'Referencia',
            'Orden/OP', 'Cant. Recibida', 'Cantidad OK', 'PNC Iny', 'PNC Pul', 'Observaciones'
        ]
        for col_idx, titulo in enumerate(columnas, 1):
            cell = ws.cell(row=1, column=col_idx, value=titulo)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # --- FILAS DE DATOS ---
        for row_idx, r in enumerate(resultados, 2):
            fecha_str = r['fecha'].strftime('%d/%m/%Y') if r['fecha'] else ''
            h_inicio = r['hora_inicio'].strftime('%H:%M') if r['hora_inicio'] else ''
            h_fin = r['hora_fin'].strftime('%H:%M') if r['hora_fin'] else ''
            cant_recibida = float(r['cantidad_recibida'] or 0)
            cant_ok = int(r['cantidad_real'] or 0)
            pnc_i = int(r['pnc_inyeccion'] or 0)
            pnc_p = int(r['pnc_pulido'] or 0)

            fila = [
                fecha_str,
                h_inicio,
                h_fin,
                str(r['responsable'] or 'SISTEMA'),
                str(r['codigo'] or ''),
                str(r['orden_produccion'] or r['id_pulido'] or '-'),
                round(cant_recibida, 1) if cant_recibida != int(cant_recibida) else int(cant_recibida),
                cant_ok,
                pnc_i,
                pnc_p,
                str(r['observaciones'] or '')
            ]

            for col_idx, valor in enumerate(fila, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                cell.border = thin_border

                # Alineación: texto a la izquierda, números al centro
                if col_idx in (4, 5, 6, 11):  # Operario, Referencia, Orden, Observaciones
                    cell.alignment = text_align
                else:
                    cell.alignment = data_align

            # Cebreado (filas pares)
            if row_idx % 2 == 0:
                for col_idx in range(1, len(columnas) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = zebra_fill

            # Alerta PNC (fondo rojo claro si > 0)
            if pnc_i > 0:
                ws.cell(row=row_idx, column=9).fill = pnc_fill
                ws.cell(row=row_idx, column=9).font = Font(bold=True, color='C0392B')
            if pnc_p > 0:
                ws.cell(row=row_idx, column=10).fill = pnc_fill
                ws.cell(row=row_idx, column=10).font = Font(bold=True, color='C0392B')

        # --- AUTO-AJUSTE DE ANCHOS ---
        anchos = [12, 11, 11, 22, 18, 18, 14, 13, 9, 9, 40]
        for i, w in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Inmovilizar primera fila
        ws.freeze_panes = 'A2'

        # 4. Generar archivo en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        fecha_archivo = datetime.now().strftime('%Y-%m-%d')
        filename = f"Historial_Pulido_{fecha_archivo}.xlsx"

        logger.info(f"📊 [Excel] Exportando {len(resultados)} registros de Pulido")

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        db.session.rollback()
        logger.error(f"Error exportando Excel Pulido: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({"success": False, "error": str(e)}), 500

# =============================================================
# NUEVO: Endpoint GET — Lotes Activos para Modo Lotes en Vivo
# Devuelve todos los lotes en estado ABIERTO_PRODUCCION.
# Sin locks: cualquier número de operarias puede consultarlo
# simultáneamente sin colisiones (solo lectura).
# =============================================================
@pulido_bp.route('/api/pulido/lotes_activos', methods=['GET'])
def get_lotes_activos():
    """
    Retorna los lotes de db_trazabilidad_lotes con estado ABIERTO_PRODUCCION.
    Pulido lo consulta al entrar al Modo Lotes en Vivo para mostrar la lista táctil.
    """
    try:
        # from sqlalchemy import or_
        # lotes = db.session.query(TrazabilidadLote).filter(
        #     or_(
        #         # Flujo normal
        #         db.and_(
        #             TrazabilidadLote.estado_actual.in_(['ABIERTO_PRODUCCION', 'ABIERTO_PULIDO']),
        #             TrazabilidadLote.por_pulir > 0
        #         ),
        #         # Forzar visibilidad para lotes de prueba
        #         TrazabilidadLote.id_lote.ilike('%PRUEBA%'),
        #         TrazabilidadLote.id_lote.ilike('%9999%'),
        #         TrazabilidadLote.orden_produccion.ilike('%PRUEBA%'),
        #         TrazabilidadLote.orden_produccion.ilike('%9999%')
        #     )
        # ).order_by(TrazabilidadLote.fecha_creacion.desc()).all()

        # Flujo estricto de producción
        lotes = db.session.query(TrazabilidadLote).filter(
            TrazabilidadLote.estado_actual == 'ABIERTO_PRODUCCION',
            TrazabilidadLote.por_pulir > 0
        ).order_by(TrazabilidadLote.fecha_creacion.desc()).all()

        resultado = [{
            'id_lote'          : l.id_lote,
            'orden_produccion' : l.orden_produccion or 'SIN OP',
            'id_codigo'        : l.id_codigo,
            'maquina'          : l.maquina or '',
            'responsable'      : l.responsable or '',
            'fecha_creacion'   : l.fecha_creacion.strftime('%d/%m/%Y %H:%M') if l.fecha_creacion else '',
            'estado_actual'    : l.estado_actual,
            'cantidad_inyectada': l.cantidad_inyectada or 0,
            'por_pulir'        : l.por_pulir or 0
        } for l in lotes]

        return jsonify({'success': True, 'lotes': resultado}), 200

    except Exception as e:
        logger.error(f'❌ [Lotes Activos] Error: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500


@pulido_bp.route('/api/pulido/liquidar_lote', methods=['POST'])
def liquidar_lote():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "No data provided"}), 400

        id_lote = data.get('id_lote')
        from backend.utils.formatters import resolver_operario
        responsable = resolver_operario(data.get('responsable')) or 'Supervisor'

        if not id_lote:
            return jsonify({"success": False, "error": "Falta id_lote"}), 400

        lote_traz = db.session.get(TrazabilidadLote, id_lote)
        if not lote_traz:
            return jsonify({"success": False, "error": f"Lote {id_lote} no encontrado"}), 404

        restante = lote_traz.por_pulir or 0
        if restante <= 0:
            lote_traz.estado_actual = 'PENDIENTE_VALIDACION'
            db.session.commit()
            return jsonify({"success": True, "message": "El lote ya estaba en 0 y fue enviado a validación"}), 200

        # Force por_pulir to 0 and change state
        lote_traz.por_pulir = 0
        lote_traz.estado_actual = 'PENDIENTE_VALIDACION'

        # Record difference as scrap in db_pulido (ProduccionPulido) and db_pnc_pulido (PncPulido)
        import uuid
        from datetime import datetime
        import pytz
        colombia_tz = pytz.timezone('America/Bogota')
        ahora = datetime.now(colombia_tz).replace(tzinfo=None)

        id_pulido = f"LIQ-{lote_traz.id_lote}"
        
        # Check if liquidation record already exists to avoid PK/unique issues if clicked twice
        existing_pulido = db.session.query(ProduccionPulido).filter_by(id_pulido=id_pulido).first()
        if not existing_pulido:
            # Create a ProduccionPulido record for the scrap
            nuevo_pulido = ProduccionPulido(
                id_pulido=id_pulido,
                fecha=ahora,
                codigo=lote_traz.id_codigo,
                responsable=responsable,
                cantidad_real=0,
                pnc_inyeccion=0,
                pnc_pulido=restante,
                hora_inicio=ahora,
                hora_fin=ahora,
                estado='FINALIZADO',
                tiempo_total_minutos=0.0,
                duracion_segundos=0,
                segundos_por_unidad=0.0,
                orden_produccion=lote_traz.orden_produccion,
                observaciones='LIQUIDACION FORZADA DE LOTE',
                criterio_pnc_pulido='LIQUIDACION FORZADA',
                lote=lote_traz.id_lote,
                cantidad_recibida=0,
                almacen_destino='P. TERMINADO'
            )
            db.session.add(nuevo_pulido)

            # Create PncPulido record
            nuevo_pnc = PncPulido(
                id_pnc_pulido=uuid.uuid4().hex[:8],
                id_pulido=id_pulido,
                codigo=lote_traz.id_codigo,
                cantidad=restante,
                criterio='LIQUIDACION FORZADA - MERMA/AJUSTE',
                codigo_ensamble='AUDITORIA PULIDO'
            )
            db.session.add(nuevo_pnc)

        db.session.commit()
        logger.info(f"⚡ [Liquidar Lote] Lote {id_lote} liquidado por {responsable}. Merma: {restante} unidades.")
        return jsonify({"success": True, "message": f"Lote {id_lote} liquidado correctamente. Merma: {restante} unidades."}), 200

    except Exception as e:
        db.session.rollback()
        logger.error(f"❌ Error liquidando lote: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@pulido_bp.route('/reporte_masivo', methods=['POST'])
@pulido_bp.route('/api/pulido/reporte_masivo', methods=['POST'])
def reporte_masivo():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "No data provided"}), 400

        hora_inicio_str = data.get('hora_inicio')
        hora_fin_str = data.get('hora_fin')
        from backend.utils.formatters import resolver_operario
        responsable = resolver_operario(data.get('responsable'))
        items = data.get('items', [])

        if not responsable or responsable == 'SISTEMA':
            return jsonify({"success": False, "error": "Falta el responsable"}), 400

        if not items:
            return jsonify({"success": False, "error": "No hay items para registrar"}), 400

        colombia_tz = pytz.timezone('America/Bogota')
        ahora = datetime.now(colombia_tz)
        fecha_actual = ahora.date()

        for item in items:
            referencia_raw = item.get('referencia')
            if not referencia_raw:
                raise ValueError("Se requiere la referencia en todos los registros del lote")

            buenos = float(item.get('buenos') or 0)
            rev_total = sum(float(r.get('cantidad', 0)) for r in item.get('revueltos', []))
            
            # Filtro de Seguridad Backend (Anti-Basura)
            if (buenos + rev_total) <= 0:
                logger.info(f"Omitiendo registro de {referencia_raw} porque la cantidad total es cero.")
                continue

            # ── Blindaje Dual: guardar con prefijo correcto en db_pulido ──
            referencia = preservar_o_normalizar_prefijo(referencia_raw)
            # referencia_sin_prefijo sólo se usa para los cruces con db_distribucion_op_pedidos
            referencia_sin_prefijo = normalizar_codigo(referencia_raw)
            op = item.get('op') or 'SIN OP'
            lote = item.get('lote') or 'SIN LOTE'

            id_pulido = f"PUL-VOZ-{uuid.uuid4().hex[:8].upper()}"
            registro = ProduccionPulido(
                id_pulido=id_pulido,
                fecha=fecha_actual,
                codigo=referencia,
                responsable=responsable,
                cantidad_real=int(buenos),
                pnc_pulido=0,
                pnc_inyeccion=0,
                orden_produccion=op,
                lote=lote,
                estado='FINALIZADO',
                departamento='PULIDO',
                cantidad_recibida=buenos,
                almacen_destino='P. TERMINADO',
                observaciones='Reporte Masivo por Voz Fin de Turno'
            )

            item_hora_inicio = item.get('hora_inicio')
            item_hora_fin = item.get('hora_fin')

            # Fallback de Tiempo
            if not item_hora_inicio or not item_hora_fin:
                registro.hora_inicio = ahora.replace(tzinfo=None)
                registro.hora_fin = ahora.replace(tzinfo=None)
                registro.duracion_segundos = 60
                registro.tiempo_total_minutos = 1.0
                registro.segundos_por_unidad = float(round(60 / buenos, 2)) if buenos > 0 else 0.0
            else:
                try:
                    hi_h, hi_m = map(int, item_hora_inicio.split(':'))
                    hf_h, hf_m = map(int, item_hora_fin.split(':'))
                    
                    t_ini = ahora.replace(hour=hi_h, minute=hi_m, second=0, microsecond=0)
                    t_fin = ahora.replace(hour=hf_h, minute=hf_m, second=0, microsecond=0)
                    
                    registro.hora_inicio = t_ini.replace(tzinfo=None)
                    
                    # Bugfix de Medianoche
                    if t_fin < t_ini:
                        from datetime import timedelta
                        t_fin += timedelta(days=1)
                    
                    registro.hora_fin = t_fin.replace(tzinfo=None)
                    
                    diff = t_fin - t_ini
                    segundos_totales = int(diff.total_seconds())
                    
                    if segundos_totales < 60:
                        segundos_totales = 60 # Contingencia mínima
                        
                    registro.duracion_segundos = segundos_totales
                    registro.tiempo_total_minutos = float(round(segundos_totales / 60.0, 2))
                    registro.segundos_por_unidad = float(round(segundos_totales / buenos, 2)) if buenos > 0 else 0.0
                    
                except Exception as e_met:
                    logger.warning(f"Error calculando duracion: {e_met}")
                    registro.hora_inicio = ahora.replace(tzinfo=None)
                    registro.hora_fin = ahora.replace(tzinfo=None)
                    registro.duracion_segundos = 60
                    registro.tiempo_total_minutos = 1.0
                    registro.segundos_por_unidad = float(round(60 / buenos, 2)) if buenos > 0 else 0.0

            db.session.add(registro)

            # --- MES PASO 2: db_productos NO se toca aquí. ---
            # La mutación de inventario ocurre EXCLUSIVAMENTE en el módulo de Validación (Paso 3).
            # Esta función solo persiste el reporte físico (db_pulido) y propaga al lote de trazabilidad.

            # --- MES PASO 2: Propagar al Lote de Trazabilidad (si viene en Modo Lotes en Vivo) ---
            id_lote_ref = item.get('id_lote')
            if id_lote_ref:
                lote_traz = db.session.get(TrazabilidadLote, id_lote_ref)
                if lote_traz:
                    # ---------------------------------------------------------
                    # Descuento en Lote Continuo (TrazabilidadLote.por_pulir)
                    # ---------------------------------------------------------
                    revueltos_items = item.get('revueltos', [])
                    rev_total = sum(float(r.get('cantidad', 0)) for r in revueltos_items)
                    procesado_lote = buenos + rev_total
                    
                    lote_traz.por_pulir = max(0, (lote_traz.por_pulir or 0) - int(procesado_lote))
                    
                    if lote_traz.por_pulir <= 0:
                        lote_traz.estado_actual = 'PENDIENTE_VALIDACION'
                        logger.info(f'🟡 [Trazabilidad] Lote {id_lote_ref} completó por_pulir -> PENDIENTE_VALIDACION por operaria {responsable}')
                    else:
                        logger.info(f'🔄 [Trazabilidad] Lote {id_lote_ref} registrado. Quedan {lote_traz.por_pulir} por pulir.')
                else:
                    logger.warning(f'⚠️ [Trazabilidad] id_lote {id_lote_ref} no encontrado en db_trazabilidad_lotes')

            # 4. Registrar Revueltos Masivos
            revueltos_items = item.get('revueltos', [])
            if revueltos_items and id_lote_ref:
                for r_item in revueltos_items:
                    r_cod_con_fr = preservar_o_normalizar_prefijo(r_item.get('id_codigo'))
                    r_cant = float(r_item.get('cantidad') or 0)
                    if r_cod_con_fr and r_cant > 0:
                        db.session.add(BujeRevuelto(
                            id_bujes_revueltos=uuid.uuid4().hex[:8],
                            id_pulido=id_pulido,
                            id_lote=id_lote_ref,
                            id_codigo=r_cod_con_fr,
                            cantidad=int(r_cant),
                            responsable=responsable
                        ))

            # --- Propagación FIFO ---
            if op and str(op).strip() != 'SIN OP':
                from backend.models.sql_models import DistribucionOpPedidos
                # ── Blindaje Dual: cruce FIFO con db_distribucion_op_pedidos usa SIN prefijo ──
                op_limpia = str(op).strip()
                codigo_limpio = referencia_sin_prefijo  # normalizado arriba, sin FR-
                
                cubetas = db.session.query(DistribucionOpPedidos).filter(
                    DistribucionOpPedidos.op_world_office == op_limpia,
                    DistribucionOpPedidos.codigo_producto == codigo_limpio
                ).order_by(DistribucionOpPedidos.id_distribucion.asc()).all()

                piezas_por_repartir = float(buenos)

                if not cubetas and piezas_por_repartir > 0:
                    pedido_asoc = db.session.query(DistribucionOpPedidos.id_pedido).filter(
                        DistribucionOpPedidos.op_world_office == op_limpia
                    ).first()
                    id_pedido_final = pedido_asoc[0] if (pedido_asoc and pedido_asoc[0]) else f"PED-IMPREVISTO-{op_limpia}"
                    
                    logger.info(f"[PULIDO-MASIVO-CONTINGENCIA] Creando cubeta temporal para OP: {op_limpia}, Producto: {codigo_limpio}")
                    nueva_cubeta = DistribucionOpPedidos(
                        op_world_office=op_limpia,
                        id_pedido=id_pedido_final,
                        codigo_producto=codigo_limpio,
                        cant_requerida=piezas_por_repartir,
                        cant_inyectada=piezas_por_repartir,
                        cant_pulida=piezas_por_repartir,
                        cant_ensamblada=0,
                        cant_alistada=0
                    )
                    db.session.add(nueva_cubeta)
                    db.session.flush()
                    cubetas = [nueva_cubeta]
                    piezas_por_repartir = 0.0

                for cubeta in cubetas:
                    if piezas_por_repartir <= 0:
                        break
                    falta = max(0, (cubeta.cant_requerida or 0) - (cubeta.cant_pulida or 0))
                    if falta > 0:
                        if piezas_por_repartir >= falta:
                            cubeta.cant_pulida = (cubeta.cant_pulida or 0) + falta
                            piezas_por_repartir -= falta
                        else:
                            cubeta.cant_pulida = (cubeta.cant_pulida or 0) + piezas_por_repartir
                            piezas_por_repartir = 0

        db.session.commit()
        return jsonify({"success": True, "message": f"Se registraron con éxito {len(items)} reportes del lote."}), 200

    except Exception as e:
        db.session.rollback()
        logger.error(f"Error en reporte_masivo: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@pulido_bp.route('/api/pnc/registrar_pulido', methods=['POST'])
def registrar_pnc_pulido():
    """
    Registra (o limpia) el desglose de PNC de Pulido en db_pnc_pulido
    y sincroniza pnc_pulido + cantidad_real en db_pulido.
    Body JSON:
      id_pulido - ID de la sesión de pulido
      id_codigo - Código del producto
      defectos  - {
          "Porosidad / Burbujas": N,
          "Marcas de Lijado / Rayones": N,
          "Desgaste Excesivo / Deformación": N,
          "Brillo Insuficiente": N
        }
    """
    try:
        data = request.get_json() or {}
        id_pulido = data.get('id_pulido')
        id_codigo = data.get('id_codigo')
        defectos  = data.get('defectos', {})

        if not id_pulido or not id_codigo:
            return jsonify({"success": False, "error": "id_pulido e id_codigo son obligatorios"}), 400

        id_cod = normalizar_codigo(id_codigo)

        # Eliminar registro previo para este turno + producto
        db.session.query(PncPulido).filter_by(id_pulido=id_pulido, codigo=id_cod).delete()

        # Mapear los 4 criterios específicos de Pulido
        porosidad = float(defectos.get("Porosidad / Burbujas", 0) or 0)
        rayones   = float(defectos.get("Marcas de Lijado / Rayones", 0) or 0)
        desgaste  = float(defectos.get("Desgaste Excesivo / Deformación", 0) or 0)
        brillo    = float(defectos.get("Brillo Insuficiente", 0) or 0)
        total_pnc = porosidad + rayones + desgaste + brillo

        prod_pul = db.session.query(ProduccionPulido).filter_by(id_pulido=id_pulido).first()

        if total_pnc > 0:
            criterio_str = (
                f"Porosidad/Burbujas: {int(porosidad)}, "
                f"Rayones: {int(rayones)}, "
                f"Desgaste/Deformación: {int(desgaste)}, "
                f"Brillo Insuficiente: {int(brillo)}"
            )
            nuevo_pnc = PncPulido(
                id_pnc_pulido=uuid.uuid4().hex[:8],
                id_pulido=id_pulido,
                codigo=id_cod,
                cantidad=total_pnc,
                criterio=criterio_str
            )
            db.session.add(nuevo_pnc)

            if prod_pul:
                prod_pul.pnc_pulido = int(round(total_pnc))
                prod_pul.criterio_pnc_pulido = criterio_str
                cant_bruta = float(prod_pul.cantidad_recibida or prod_pul.cantidad_real or 0)
                prod_pul.cantidad_real = max(0, int(round(cant_bruta - total_pnc)))

            db.session.commit()
            logger.info(f"✅ PNC Pulido registrado para {id_cod} en {id_pulido}: Total={total_pnc}")
            return jsonify({
                "success": True,
                "message": "PNC de Pulido registrado en db_pnc_pulido",
                "total_pnc": total_pnc
            }), 200
        else:
            if prod_pul:
                prod_pul.pnc_pulido = 0
                prod_pul.criterio_pnc_pulido = ""
            db.session.commit()
            return jsonify({
                "success": True,
                "message": "Sin defectos de PNC para Pulido",
                "total_pnc": 0
            }), 200

    except Exception as e:
        db.session.rollback()
        logger.error(f"❌ Error registrando PNC Pulido: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
