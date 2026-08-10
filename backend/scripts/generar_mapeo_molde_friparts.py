# -*- coding: utf-8 -*-
"""
Generador de Mapeo Molde <-> Codigo Friparts
---------------------------------------------
NO modifica la base de datos ni el repo real (db_moldes, db_productos). Cruza
archivos que el usuario entrego manualmente (inventario fisico de molderia) y
produce un Excel de mapeo listo para revision, siguiendo el mismo patron que
exportar_candidatos_correccion_tiempo.py: generar candidatos + cola de
revision manual, nunca inferir el dato faltante.

Fuentes (copiadas a scratch/moldes/):
  - INVENTARIO_MOLDERIA_normalizado.xlsx (entregado 2026-08-05): salida ya
    normalizada a 3FN de un inventario manual de molderia hecho la semana del
    2026-07-27 (345 filas originales). Usa rel_molde_referencia (que
    referencia produce cada molde, con cuantas cavidades) y
    rel_molde_moneda_alternativa (monedas intercambiables). Trae ya resueltas
    las reglas de split de codigos combinados "A-B" y la clasificacion
    tipo_entidad (MOLDE_COMPLETO / MACHO_INDEPENDIENTE / HERRAMIENTA_MANUAL /
    MOLDE_MONEDA).
  - "INVENTARIO MOLDERIA 28_07_2026.xlsx" (entregado 2026-08-05, reemplaza una
    version anterior que el usuario senalo como desactualizada): fuente cruda
    equivalente pero mas reciente. Se usa SOLO para detectar codigos nuevos
    que el archivo normalizado no tenia (diff automatico por codigo, ver
    _codigos_nuevos_no_normalizados) y para extraer el inventario de machos
    (que el archivo normalizado no expone como entidad propia).
  - "Base de datos ext e int.xlsx": hoja BASE, geometria por Codigo Friparts
    (OEM, DIAMETRO INTERNO/EXTERNO, ALTURA, PARED, DESCRIPCION).

Cruce: codigo_referencia (moldes) contra Codigo Friparts (geometria), como
string exacto — sin prefijos que inferir (ver feedback_prefijos_codigo_producto:
nunca se asume division FR-/MT-/CAR- por formato numerico, aqui tampoco).

Machos (2026-08-05, aclaracion del usuario): NO se excluyen del sistema, solo
del cruce Friparts (no tienen SKU propio vendible). Son una restriccion real
de planta — si una referencia necesita montar 9 machos "19" y solo hay 8
fisicos, no se puede programar esa cantidad. Van en su propia hoja
'inventario_machos' con la cantidad fisica disponible (columna "Numero
Cavidades" del inventario crudo, reutilizada aqui como conteo de unidades
fisicas del macho, no como cavidades de molde). Regla de negocio pendiente de
aplicar (falta el dato de tamano/diametro de cada macho, no esta en ningun
archivo entregado): tolerancia de +/-1.5mm al emparejar un macho con el
diametro que requiere una referencia.

Codigos confirmados por el usuario como existentes pero sin geometria
cargada: quedan en 'mapeo_valido' (el vinculo molde-referencia SI esta
confirmado) con las columnas geometricas en null — no se excluyen ni se
inventan medidas.

'9995' queda tal cual en cola_revision, sin resolver — el usuario indico que
es un caso distinto que va a aclarar aparte.

Ejecutar manualmente:
    python backend/scripts/generar_mapeo_molde_friparts.py
"""

import os
import re

import pandas as pd
import openpyxl

BASE_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'scratch', 'moldes')
PATH_MOLDERIA_NORMALIZADO = os.path.join(BASE_DIR, 'INVENTARIO_MOLDERIA_normalizado.xlsx')
PATH_MOLDERIA_CRUDO = os.path.join(BASE_DIR, 'INVENTARIO MOLDERIA 28_07_2026.xlsx')
PATH_GEOMETRIA = os.path.join(BASE_DIR, 'Base de datos ext e int.xlsx')
SALIDA_XLSX = os.path.join(BASE_DIR, 'rel_molde_friparts_mapeo_v9.xlsx')

TIPOS_SIN_SKU_PROPIO = {'HERRAMIENTA_MANUAL', 'MACHO_INDEPENDIENTE'}

# Crosswalk confirmado por el usuario (2026-08-05): codigo interno de molderia
# -> Codigo Friparts real. Corroborado por la propia base geometrica: las filas
# destino traen "RB-410"/"RB-413"/"RB-440" en OEM/DESCRIPCION, o sea el codigo
# viejo quedo registrado ahi mismo.
CROSSWALK_CONFIRMADO = {
    '410': '9609',
    '413': '9722',
    '440': '9723',
}

# Codigos confirmados por el usuario (2026-08-05) como moldes/referencias que
# SI existen fisicamente pero no tienen entrada en la base geometrica todavia.
CONFIRMADOS_FALTA_CATALOGO = {'9076', '9997', '9999', '7034', '5002C', '5015', '9881v', '9881b', '9992', '5012'}

# Referencias nuevas confirmadas manualmente por el usuario, que no vienen en
# ningun inventario de molderia todavia (se agregaron en planta el dia
# anterior a la confirmacion).
REFERENCIAS_NUEVAS_CONFIRMADAS = [
    {
        'origen': 'confirmado_manual_2026-08-06',
        'molde_id': None,
        'codigo_molde_original': 'D',
        'codigo_referencia': '5012',
        'cavidades': 1,
        'tipo_vinculo': 'CAVIDAD_FIJA',
        'tipo_entidad': 'MOLDE_COMPLETO',
    },
    # Molde nuevo en pruebas (2026-08-06), 32 cavidades combo montado en
    # Maquina 1 / portamolde P. Sin codigo definitivo de planta todavia; el
    # usuario confirmo usar este identificador temporal. Las referencias
    # 9630/9316 ya tienen su molde establecido (3/2 cavidades, M/N/Ñ) - este
    # es un SEGUNDO molde distinto que produce las mismas referencias.
    {
        'origen': 'confirmado_manual_2026-08-06_molde_prueba',
        'molde_id': None,
        'codigo_molde_original': 'P-PRUEBA-9630-9316',
        'codigo_referencia': '9630',
        'cavidades': 20,
        'tipo_vinculo': 'CAVIDAD_FIJA',
        'tipo_entidad': 'MOLDE_COMPLETO',
    },
    {
        'origen': 'confirmado_manual_2026-08-06_molde_prueba',
        'molde_id': None,
        'codigo_molde_original': 'P-PRUEBA-9630-9316',
        'codigo_referencia': '9316',
        'cavidades': 12,
        'tipo_vinculo': 'CAVIDAD_FIJA',
        'tipo_entidad': 'MOLDE_COMPLETO',
    },
]

# Confirmado por el usuario (2026-08-06): son tolerancias distintas, no la
# misma. TOLERANCIA_MACHO_MM se aplica al calzar el diametro de un macho
# contra el diametro que pide una referencia. TOLERANCIA_PARED_MM se aplica
# (a futuro, todavia no hay logica que la use) al agrupar referencias por
# PARED como proxy de tiempo de ciclo compatible.
TOLERANCIA_MACHO_MM = 1.0
TOLERANCIA_PARED_MM = 1.5

# DESCARTADO (2026-08-06): esta division "maquinas grandes 1,4 = M,N,Ñ,O,P /
# chicas 2,3 = A-M" se cruzo contra el historico real de db_inyeccion y no se
# sostiene — 54 de 543 combinaciones maquina-referencia solo eran posibles
# con un portamolde fuera del grupo "asignado", y Maquina 4 en particular
# corrio moldes de casi todo el rango chico (A,B,C,D,E,F,I). Ya NO se carga a
# rel_maquina_portamolde (ver cargar_moldes_portamoldes_machos.py). Se deja
# el diccionario aqui solo como referencia historica de la hipotesis
# descartada; no usar para poblar la tabla real hasta tener el patron
# correcto (pendiente que el usuario muestre ejemplos reales de montajes).
MAQUINAS_PORTAMOLDES = {
    'Maquina 1': {'capacidad': 'mayor', 'portamoldes': ['M', 'N', 'Ñ', 'O', 'P']},
    'Maquina 4': {'capacidad': 'mayor', 'portamoldes': ['M', 'N', 'Ñ', 'O', 'P']},
    'Maquina 2': {'capacidad': 'menor', 'portamoldes': ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']},
    'Maquina 3': {'capacidad': 'menor', 'portamoldes': ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']},
}

VALORES_ERROR_EXCEL = {'#VALUE!', '#REF!', '#DIV/0!', '#N/A', '#NAME?', '#NULL!', '#NUM!', '#¡VALOR!'}


def _limpiar_error_excel(valor):
    """PARED en 'Base de datos ext e int.xlsx' es una formula '=(EXT-INT)/2'
    que Excel ya trae rota (#VALUE!) en 63 filas de origen, porque INTERNO o
    EXTERNO no es un numero limpio ahi (pulgadas, roscas, rangos con coma).
    No se inventa el numero: se limpia el artefacto de error de Excel a None
    y se deja constancia en 'pared_formula_rota_en_origen' para que quede
    visible, sin tocar ØINTERNO/ØEXTERNO crudos (esa limpieza es una decision
    de equipo de molderia, no de este script)."""
    if isinstance(valor, str) and valor.strip().upper() in VALORES_ERROR_EXCEL:
        return None
    return valor


def _norm_codigo(c):
    if c is None:
        return None
    if isinstance(c, float) and c.is_integer():
        return str(int(c))
    return str(c).strip()


def _cargar_moldes_normalizados():
    wb = openpyxl.load_workbook(PATH_MOLDERIA_NORMALIZADO, data_only=True)

    ws_moldes = wb['db_moldes']
    tipo_por_molde_id = {}
    codigos_normalizados = set()
    for _id, codigo_original, tipo_entidad, fam_id, fam_nombre, fila_origen in ws_moldes.iter_rows(min_row=2, values_only=True):
        tipo_por_molde_id[_id] = tipo_entidad
        codigos_normalizados.add(_norm_codigo(codigo_original))

    ws_rel = wb['rel_molde_referencia']
    referencias = []
    for molde_id, codigo_molde_original, codigo_referencia, cavidades, tipo_vinculo in ws_rel.iter_rows(min_row=2, values_only=True):
        if codigo_referencia is None:
            continue
        referencias.append({
            'origen': 'rel_molde_referencia',
            'molde_id': molde_id,
            'codigo_molde_original': codigo_molde_original,
            'codigo_referencia': str(codigo_referencia).strip(),
            'cavidades': cavidades,
            'tipo_vinculo': tipo_vinculo,
            'tipo_entidad': tipo_por_molde_id.get(molde_id, '?'),
        })

    ws_moneda = wb['rel_molde_moneda_alternativa']
    for molde_id, codigo_molde_original, codigo_referencia_alt, cantidad_moneda in ws_moneda.iter_rows(min_row=2, values_only=True):
        if codigo_referencia_alt is None:
            continue
        referencias.append({
            'origen': 'rel_molde_moneda_alternativa',
            'molde_id': molde_id,
            'codigo_molde_original': codigo_molde_original,
            'codigo_referencia': str(codigo_referencia_alt).strip(),
            'cavidades': cantidad_moneda,
            'tipo_vinculo': 'MONEDA_ALTERNATIVA',
            'tipo_entidad': tipo_por_molde_id.get(molde_id, '?'),
        })

    return referencias, codigos_normalizados


def _cargar_crudo():
    wb = openpyxl.load_workbook(PATH_MOLDERIA_CRUDO, data_only=True)
    ws = wb['Hoja 1']
    filas = []
    for codigo, num_cavidades, familia, portamoldes, *_resto in ws.iter_rows(min_row=2, values_only=True):
        if codigo is None:
            continue
        filas.append({
            'codigo_raw': codigo,
            'codigo': _norm_codigo(codigo),
            'numero_cavidades_raw': num_cavidades,
            'familia': familia,
            'portamoldes': portamoldes,
        })
    return filas


def _es_macho(codigo_raw):
    return isinstance(codigo_raw, str) and codigo_raw.strip().lower().startswith('macho')


def _codigos_nuevos_no_normalizados(filas_crudo, codigos_normalizados):
    """Codigos que aparecen en el inventario crudo mas reciente pero no en el
    normalizado (345 filas). Se agregan como MOLDE_COMPLETO/CAVIDAD_FIJA solo
    si son codigos simples (sin combos 'A-B' ni punto) - un codigo combinado
    nuevo requeriria la misma revision manual que ya se aplico a los del
    archivo normalizado, y no se va a replicar esa heuristica aqui."""
    nuevos = []
    for fila in filas_crudo:
        codigo = fila['codigo']
        if codigo in codigos_normalizados or _es_macho(fila['codigo_raw']):
            continue
        es_combo = bool(re.search(r'[-.]', codigo)) if not codigo.replace('.', '').isdigit() else False
        if '-' in str(fila['codigo_raw']) or (isinstance(fila['codigo_raw'], str) and '.' in fila['codigo_raw']):
            es_combo = True
        nuevos.append({
            'origen': 'crudo_28_07_2026_codigo_no_visto',
            'molde_id': None,
            'codigo_molde_original': fila['codigo_raw'],
            'codigo_referencia': codigo,
            'cavidades': fila['numero_cavidades_raw'],
            'tipo_vinculo': 'CAVIDAD_FIJA' if not es_combo else 'COMBO_SIN_NORMALIZAR',
            'tipo_entidad': 'MOLDE_COMPLETO' if not es_combo else '?',
        })
    return nuevos


def _cargar_inventario_machos(filas_crudo):
    """El diametro interno del macho es el numero en su propio nombre (ej.
    'Macho 19' -> 19mm interno), confirmado por el usuario 2026-08-05. No
    aplica a 'Macho suavizador cavidad' (sin numero) - queda en null."""
    machos = []
    for fila in filas_crudo:
        if not _es_macho(fila['codigo_raw']):
            continue
        nombre = fila['codigo_raw'].strip()
        match = _RE_MACHO_NUMERO.match(nombre)
        machos.append({
            'codigo_macho': nombre,
            'diametro_interno_mm': int(match.group(1)) if match else None,
            'cantidad_fisica_disponible': fila['numero_cavidades_raw'],
            'portamoldes': fila['portamoldes'],
            'tolerancia_mm': TOLERANCIA_MACHO_MM,
        })
    return machos


def _cargar_geometria():
    wb = openpyxl.load_workbook(PATH_GEOMETRIA, data_only=True)
    ws = wb['BASE']
    por_codigo = {}
    for codigo_friparts, oem, interno, externo, altura, pared, descripcion in ws.iter_rows(min_row=2, values_only=True):
        if codigo_friparts is None:
            continue
        pared_limpio = _limpiar_error_excel(pared)
        # (vacio - vacio) / 2 = 0 en Excel: un 0 sin ØINTERNO/ØEXTERNO reales
        # no es un grosor de pared valido, es la misma ausencia de dato que el
        # #VALUE!, solo que sin verse como error.
        if interno is None and externo is None:
            pared_limpio = None
        por_codigo[str(codigo_friparts).strip()] = {
            'codigo_friparts': str(codigo_friparts).strip(),
            'oem': oem,
            'diametro_interno': interno,
            'diametro_externo': externo,
            'altura': altura,
            'pared': pared_limpio,
            'pared_formula_rota_en_origen': pared_limpio is None and pared is not None,
            'descripcion': descripcion,
        }
    return por_codigo


GEOMETRIA_VACIA = {
    'codigo_friparts': None, 'oem': None, 'diametro_interno': None,
    'diametro_externo': None, 'altura': None, 'pared': None,
    'pared_formula_rota_en_origen': False, 'descripcion': None,
}

_RE_MACHO_NUMERO = re.compile(r'macho\s*(\d+)', re.IGNORECASE)


def main():
    referencias_normalizadas, codigos_normalizados = _cargar_moldes_normalizados()
    filas_crudo = _cargar_crudo()
    referencias_nuevas = _codigos_nuevos_no_normalizados(filas_crudo, codigos_normalizados)
    inventario_machos = _cargar_inventario_machos(filas_crudo)
    geometria = _cargar_geometria()

    referencias = referencias_normalizadas + referencias_nuevas + REFERENCIAS_NUEVAS_CONFIRMADAS

    validas, sin_resolver = [], []
    for fila in referencias:
        if fila['tipo_vinculo'] == 'COMBO_SIN_NORMALIZAR':
            sin_resolver.append({**fila, 'motivo': 'CODIGO_NUEVO_COMBINADO: aparece en el inventario 28/07 pero no en el normalizado 3FN, y trae guion/punto - necesita la misma revision manual que los combos ya resueltos, no se debe adivinar el split de cavidades.'})
            continue
        if fila['tipo_entidad'] in TIPOS_SIN_SKU_PROPIO:
            continue  # va en inventario_machos / no aplica geometria Friparts

        codigo = fila['codigo_referencia']
        codigo_resuelto = CROSSWALK_CONFIRMADO.get(codigo, codigo)
        geo = geometria.get(codigo_resuelto)

        if geo:
            validas.append({**fila, 'codigo_referencia_original': codigo, 'via_crosswalk': codigo != codigo_resuelto, 'geometria_pendiente': False, **geo})
        elif codigo in CONFIRMADOS_FALTA_CATALOGO:
            validas.append({**fila, 'codigo_referencia_original': codigo, 'via_crosswalk': False, 'geometria_pendiente': True, **GEOMETRIA_VACIA})
        else:
            sin_resolver.append({**fila, 'motivo': 'SIN_RESOLVER: no aparece en la base geometrica ni fue confirmado como falta de catalogo (ej. 9995, ambiguo por duplicado de familia).'})

    df_validas = pd.DataFrame(validas)
    df_sin_resolver = pd.DataFrame(sin_resolver)
    df_machos = pd.DataFrame(inventario_machos)
    df_maquinas = pd.DataFrame([
        {'maquina': maquina, 'capacidad': info['capacidad'], 'portamolde': portamolde, 'cantidad_fisica': 1}
        for maquina, info in MAQUINAS_PORTAMOLDES.items()
        for portamolde in info['portamoldes']
    ])

    resumen = pd.DataFrame([
        {'metrica': 'total_referencias_molde (normalizado + nuevas del 28_07)', 'valor': len(referencias)},
        {'metrica': 'mapeadas_a_friparts_con_geometria', 'valor': sum(1 for v in validas if not v['geometria_pendiente'])},
        {'metrica': 'mapeadas_via_crosswalk_confirmado', 'valor': sum(1 for v in validas if v['via_crosswalk'])},
        {'metrica': 'mapeadas_geometria_pendiente_confirmada', 'valor': sum(1 for v in validas if v['geometria_pendiente'])},
        {'metrica': 'sin_resolver', 'valor': len(sin_resolver)},
        {'metrica': 'machos_en_inventario_aparte (no cuentan como gap)', 'valor': len(inventario_machos)},
        {'metrica': 'geometria_con_pared_formula_rota_en_origen', 'valor': sum(1 for v in validas if v.get('pared_formula_rota_en_origen'))},
        {'metrica': 'tolerancia_macho_mm', 'valor': TOLERANCIA_MACHO_MM},
        {'metrica': 'tolerancia_pared_mm (pendiente de aplicar en logica)', 'valor': TOLERANCIA_PARED_MM},
    ])

    with pd.ExcelWriter(SALIDA_XLSX, engine='openpyxl') as writer:
        df_validas.to_excel(writer, sheet_name='mapeo_valido', index=False)
        df_sin_resolver.to_excel(writer, sheet_name='sin_resolver', index=False)
        df_machos.to_excel(writer, sheet_name='inventario_machos', index=False)
        df_maquinas.to_excel(writer, sheet_name='maquinas_portamoldes', index=False)
        resumen.to_excel(writer, sheet_name='resumen', index=False)

    print(f"Generado: {SALIDA_XLSX}")
    print(resumen.to_string(index=False))


if __name__ == '__main__':
    main()
