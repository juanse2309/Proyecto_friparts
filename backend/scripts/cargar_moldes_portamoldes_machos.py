# -*- coding: utf-8 -*-
"""
Carga a la base de datos real: rel_producto_molde, db_portamoldes,
rel_molde_portamoldes, rel_maquina_portamolde, db_machos, y las 4 columnas
geometricas de db_productos.

Reutiliza integramente la logica de cruce/limpieza de
generar_mapeo_molde_friparts.py (mismo dataset ya revisado en
rel_molde_friparts_mapeo_v8.xlsx) — no se recalcula ni se reinterpreta nada
aparte, lo que se carga en la BD es exactamente lo que ya se te mostro ahi.

Reglas de carga:
  - rel_producto_molde, rel_molde_portamoldes, rel_maquina_portamolde: se
    vacian y se recargan completas en cada corrida (DELETE + INSERT). Son
    propiedad exclusiva de este script -> idempotente, correr de nuevo no
    duplica filas.
  - db_portamoldes y db_machos: UPSERT por su codigo unico (ON CONFLICT).
  - db_productos.diametro_interno/externo/altura/pared: UPDATE por
    id_codigo = Codigo Friparts EXACTO, sin agregar ni quitar prefijo (ver
    feedback_prefijos_codigo_producto — id_codigo en la BD real ya vive sin
    'FR-', igual que la columna "Codigo Friparts" del Excel de geometria).
    Se actualizan las 441 referencias de la base geometrica completa, no
    solo las 348 que ya tienen molde asignado. Valores de texto no numerico
    (pulgadas, roscas, rangos — ver auditoria previa) se dejan en NULL, nunca
    se intenta convertir/adivinar el numero.
  - rel_molde_portamoldes agrega a mano los 2 moldes nuevos del inventario
    28/07 (9913, 9953, portamoldes 'A B' tal como vienen en esa fila) y
    '5012 -> D' (el molde de esa referencia nueva ES el portamolde D, per
    confirmacion del usuario 2026-08-06).
  - '9995' (sin_resolver) NO se carga en rel_producto_molde — sigue
    pendiente de que el usuario confirme si son moldes fisicos distintos.

Ejecutar manualmente (desde la raiz del repo, por los imports absolutos):
    python -m backend.scripts.cargar_moldes_portamoldes_machos
"""
import openpyxl
from sqlalchemy import text

from backend.app import app
from backend.core.sql_database import db
from backend.scripts.generar_mapeo_molde_friparts import (
    _cargar_moldes_normalizados, _cargar_crudo, _codigos_nuevos_no_normalizados,
    _cargar_inventario_machos, _cargar_geometria, _norm_codigo,
    CROSSWALK_CONFIRMADO, CONFIRMADOS_FALTA_CATALOGO, REFERENCIAS_NUEVAS_CONFIRMADAS,
    TIPOS_SIN_SKU_PROPIO, MAQUINAS_PORTAMOLDES, GEOMETRIA_VACIA,
    PATH_MOLDERIA_NORMALIZADO,
)


def _construir_mapeo():
    """Reconstruye exactamente el mismo dataset 'validas' que genera
    generar_mapeo_molde_friparts.main() para la hoja mapeo_valido, sin volver
    a escribir el Excel (esa parte ya se hizo y se revisó)."""
    referencias_normalizadas, codigos_normalizados = _cargar_moldes_normalizados()
    filas_crudo = _cargar_crudo()
    referencias_nuevas = _codigos_nuevos_no_normalizados(filas_crudo, codigos_normalizados)
    inventario_machos = _cargar_inventario_machos(filas_crudo)
    geometria = _cargar_geometria()

    referencias = referencias_normalizadas + referencias_nuevas + REFERENCIAS_NUEVAS_CONFIRMADAS

    validas = []
    for fila in referencias:
        if fila['tipo_vinculo'] == 'COMBO_SIN_NORMALIZAR':
            continue
        if fila['tipo_entidad'] in TIPOS_SIN_SKU_PROPIO:
            continue
        codigo = fila['codigo_referencia']
        codigo_resuelto = CROSSWALK_CONFIRMADO.get(codigo, codigo)
        geo = geometria.get(codigo_resuelto)
        if geo:
            validas.append({**fila, 'geometria_pendiente': False, **geo})
        elif codigo in CONFIRMADOS_FALTA_CATALOGO:
            validas.append({**fila, 'geometria_pendiente': True, **GEOMETRIA_VACIA})
        # sin_resolver (9995) queda fuera a proposito.

    return validas, inventario_machos, geometria, filas_crudo


def _construir_rel_molde_portamoldes(filas_crudo):
    """(codigo_molde, codigo_portamolde) para cada molde real (excluye machos
    y herramientas — esos no tienen 'cuerpo de molde' que montar)."""
    wb = openpyxl.load_workbook(PATH_MOLDERIA_NORMALIZADO, data_only=True)
    ws_moldes = wb['db_moldes']
    tipo_por_id = {}
    for _id, codigo_original, tipo_entidad, fam_id, fam_nombre, fila in ws_moldes.iter_rows(min_row=2, values_only=True):
        tipo_por_id[_id] = tipo_entidad

    ws_port = wb['rel_molde_portamoldes']
    pares = []
    for molde_id, codigo_molde_original, portamolde_id, portamolde_codigo in ws_port.iter_rows(min_row=2, values_only=True):
        if tipo_por_id.get(molde_id) in TIPOS_SIN_SKU_PROPIO:
            continue
        pares.append((_norm_codigo(codigo_molde_original), portamolde_codigo))

    for fila in filas_crudo:
        if fila['codigo'] in ('9913', '9953') and fila['portamoldes']:
            for p in str(fila['portamoldes']).split():
                pares.append((fila['codigo'], p))

    pares.append(('D', 'D'))  # 5012 -> molde D, confirmado por el usuario 2026-08-06
    pares.append(('P-PRUEBA-9630-9316', 'P'))  # molde nuevo en pruebas, confirmado 2026-08-06
    return pares


def cargar():
    validas, inventario_machos, geometria, filas_crudo = _construir_mapeo()
    pares_molde_portamolde = _construir_rel_molde_portamoldes(filas_crudo)
    todas_las_letras = sorted({p for info in MAQUINAS_PORTAMOLDES.values() for p in info['portamoldes']})

    with app.app_context():
        try:
            for letra in todas_las_letras:
                db.session.execute(text("""
                    INSERT INTO db_portamoldes (codigo, cantidad_fisica, activo)
                    VALUES (:codigo, 1, TRUE)
                    ON CONFLICT (codigo) DO UPDATE SET cantidad_fisica = 1, activo = TRUE
                """), {'codigo': letra})

            db.session.execute(text("DELETE FROM rel_maquina_portamolde"))
            n_rel_maquina = 0
            for maquina, info in MAQUINAS_PORTAMOLDES.items():
                for p in info['portamoldes']:
                    db.session.execute(text("""
                        INSERT INTO rel_maquina_portamolde (maquina, codigo_portamolde)
                        VALUES (:maquina, :p)
                    """), {'maquina': maquina, 'p': p})
                    n_rel_maquina += 1

            for m in inventario_machos:
                db.session.execute(text("""
                    INSERT INTO db_machos (codigo_macho, diametro_interno_mm, cantidad_fisica_disponible, activo)
                    VALUES (:codigo, :diam, :cant, TRUE)
                    ON CONFLICT (codigo_macho) DO UPDATE SET
                        diametro_interno_mm = EXCLUDED.diametro_interno_mm,
                        cantidad_fisica_disponible = EXCLUDED.cantidad_fisica_disponible,
                        activo = TRUE
                """), {
                    'codigo': m['codigo_macho'],
                    'diam': m['diametro_interno_mm'],
                    'cant': m['cantidad_fisica_disponible'],
                })

            db.session.execute(text("DELETE FROM rel_producto_molde"))
            for v in validas:
                ref = v.get('codigo_friparts') or _norm_codigo(v['codigo_referencia'])
                db.session.execute(text("""
                    INSERT INTO rel_producto_molde (codigo_molde, codigo_referencia, cavidades, tipo_vinculo, activo)
                    VALUES (:molde, :ref, :cav, :tipo, TRUE)
                """), {
                    'molde': _norm_codigo(v['codigo_molde_original']),
                    'ref': ref,
                    'cav': int(v['cavidades']) if v['cavidades'] is not None else 1,
                    'tipo': v['tipo_vinculo'],
                })

            db.session.execute(text("DELETE FROM rel_molde_portamoldes"))
            for codigo_molde, codigo_portamolde in pares_molde_portamolde:
                db.session.execute(text("""
                    INSERT INTO rel_molde_portamoldes (codigo_molde, codigo_portamolde)
                    VALUES (:m, :p)
                """), {'m': codigo_molde, 'p': codigo_portamolde})

            actualizados = 0
            for codigo, geo in geometria.items():
                def _num(x):
                    return x if isinstance(x, (int, float)) else None
                result = db.session.execute(text("""
                    UPDATE db_productos
                    SET diametro_interno = :di, diametro_externo = :de, altura = :al, pared = :pa
                    WHERE id_codigo = :codigo
                """), {
                    'di': _num(geo['diametro_interno']),
                    'de': _num(geo['diametro_externo']),
                    'al': _num(geo['altura']),
                    'pa': geo['pared'],
                    'codigo': codigo,
                })
                actualizados += result.rowcount

            db.session.commit()
            print("Carga exitosa:")
            print(f"  db_portamoldes: {len(todas_las_letras)} upserted")
            print(f"  rel_maquina_portamolde: {n_rel_maquina} filas")
            print(f"  db_machos: {len(inventario_machos)} upserted")
            print(f"  rel_producto_molde: {len(validas)} filas")
            print(f"  rel_molde_portamoldes: {len(pares_molde_portamolde)} filas")
            print(f"  db_productos con geometria actualizada: {actualizados} de {len(geometria)} codigos en la base geometrica")
        except Exception as e:
            db.session.rollback()
            print("Error en carga:", e)
            raise


if __name__ == '__main__':
    cargar()
