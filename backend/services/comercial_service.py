import logging
import re
from datetime import datetime, date
from sqlalchemy import text
from backend.core.sql_database import db

logger = logging.getLogger(__name__)

# Aislamiento estricto por igualdad exacta (no LIKE): permite uso de indice B-Tree
# sobre db_ventas.vendedor y evita falsos positivos entre nombres similares.
FILTRO_SCOPE_ROL = "(:es_comercial = FALSE OR UPPER(TRIM(COALESCE(v.vendedor, ''))) = UPPER(TRIM(:vendedor_user)))"
FILTRO_VENDEDOR_OPCIONAL = "(:vendedor_filtro = '' OR UPPER(TRIM(COALESCE(v.vendedor, ''))) = UPPER(TRIM(:vendedor_filtro)))"


class ComercialHistoricoService:
    """
    Servicio de extracción y analítica comercial histórica.
    Lee vendedor y zona de forma plana desde las columnas nativas de db_ventas
    (pobladas por World Office: Nombres_tercero_interno / Ciudad_Encabezado).
    Garantiza aislamiento estricto por rol y fidelidad contable restando
    devoluciones (Notas Crédito) directamente en PostgreSQL.
    """

    @staticmethod
    def _expr_ajustado_nc(columna: str) -> str:
        """CASE que invierte el signo de `columna` cuando el registro es una Nota Crédito."""
        return f"""
            CASE
                WHEN UPPER(TRIM(COALESCE(v.clasificacion, ''))) LIKE '%NC%'
                  OR UPPER(TRIM(COALESCE(v.documento, ''))) LIKE '%NC%'
                THEN -ABS(COALESCE(v.{columna}, 0))
                ELSE COALESCE(v.{columna}, 0)
            END
        """

    @staticmethod
    def _resolver_alias_vendedor(username: str) -> str:
        """
        Resuelve el valor a comparar contra db_ventas.vendedor (match exacto):
        prioriza db_usuarios.alias_vendedor_wo (nombre tal como lo escribe World
        Office) y cae al nombre de login si el alias no está configurado.

        Se busca por `username` (no por user_id de sesión): ningún endpoint de
        login en auth_routes.py escribe session['user_id']/['usuario_id'], por lo
        que ese id siempre llega en 0 y resolvería, para cualquier usuario, el
        alias del registro id=0 en vez del propio.
        """
        try:
            from backend.models.sql_models import Usuario
            if username:
                usuario = Usuario.query.filter_by(username=str(username)).first()
                if usuario and usuario.alias_vendedor_wo and usuario.alias_vendedor_wo.strip():
                    return usuario.alias_vendedor_wo.strip()
        except Exception as e:
            logger.warning(f"[COMERCIAL_SERVICE] No se pudo resolver alias_vendedor_wo para username={username!r}: {e}")
        return str(username or '')

    @staticmethod
    def _resolver_fecha_corte(fecha_corte_raw) -> date:
        """
        Valida y convierte fecha_corte a un date real. Si no viene o es invalida,
        usa hoy. Nunca deja pasar un string sin validar hacia la consulta SQL.
        """
        if fecha_corte_raw:
            try:
                return datetime.strptime(str(fecha_corte_raw)[:10], '%Y-%m-%d').date()
            except (ValueError, TypeError):
                logger.warning(f"[COMERCIAL_SERVICE] fecha_corte inválida recibida: {fecha_corte_raw!r}. Usando hoy.")
        return datetime.now().date()

    @staticmethod
    def obtener_analitica_historica(user_id: int, username: str, user_role: str, start_year: int = 2024, end_year: int = 2026) -> dict:
        """
        Extrae datos consolidados de ventas (Año, Mes, Zona, Cliente) entre start_year y end_year.
        Resta devoluciones/NC. Vendedor y zona se leen directo de db_ventas (sin JOIN).
        """
        role_upper = str(user_role or '').strip().upper()
        es_global = role_upper in ['ADMIN', 'ADMINISTRACION', 'ADMINISTRADOR', 'GERENCIA']
        es_comercial = not es_global

        ventas_expr = ComercialHistoricoService._expr_ajustado_nc('total_ingresos')
        cantidad_expr = ComercialHistoricoService._expr_ajustado_nc('cantidad')
        vendedor_scope = ComercialHistoricoService._resolver_alias_vendedor(username)

        params = {
            'start_year': int(start_year),
            'end_year': int(end_year),
            'es_comercial': es_comercial,
            'vendedor_user': vendedor_scope,
            'vendedor_filtro': ''  # sin acotar adicionalmente en esta vista
        }

        query_resumen = text(f"""
            SELECT
                EXTRACT(YEAR FROM v.fecha)::INTEGER AS anio,
                ROUND(SUM({ventas_expr})::NUMERIC, 2) AS total_ventas,
                ROUND(SUM({cantidad_expr})::NUMERIC, 2) AS total_unidades,
                COUNT(v.id)::INTEGER AS total_transacciones
            FROM db_ventas v
            WHERE v.fecha >= MAKE_DATE(:start_year, 1, 1)
              AND v.fecha <= MAKE_DATE(:end_year, 12, 31)
              AND {FILTRO_SCOPE_ROL}
              AND {FILTRO_VENDEDOR_OPCIONAL}
            GROUP BY EXTRACT(YEAR FROM v.fecha)
            ORDER BY anio ASC;
        """)

        query_zonas = text(f"""
            SELECT
                COALESCE(NULLIF(TRIM(v.zona), ''), 'SIN ZONA') AS zona,
                ROUND(SUM({ventas_expr})::NUMERIC, 2) AS total_ventas,
                ROUND(SUM({cantidad_expr})::NUMERIC, 2) AS total_unidades
            FROM db_ventas v
            WHERE v.fecha >= MAKE_DATE(:start_year, 1, 1)
              AND v.fecha <= MAKE_DATE(:end_year, 12, 31)
              AND {FILTRO_SCOPE_ROL}
              AND {FILTRO_VENDEDOR_OPCIONAL}
            GROUP BY COALESCE(NULLIF(TRIM(v.zona), ''), 'SIN ZONA')
            ORDER BY total_ventas DESC;
        """)

        query_top_clientes = text(f"""
            SELECT
                COALESCE(NULLIF(TRIM(v.nombres), ''), 'CLIENTE DESCONOCIDO') AS cliente,
                ROUND(SUM({ventas_expr})::NUMERIC, 2) AS total_ventas,
                ROUND(SUM({cantidad_expr})::NUMERIC, 2) AS total_unidades
            FROM db_ventas v
            WHERE v.fecha >= MAKE_DATE(:start_year, 1, 1)
              AND v.fecha <= MAKE_DATE(:end_year, 12, 31)
              AND {FILTRO_SCOPE_ROL}
              AND {FILTRO_VENDEDOR_OPCIONAL}
            GROUP BY COALESCE(NULLIF(TRIM(v.nombres), ''), 'CLIENTE DESCONOCIDO')
            ORDER BY total_ventas DESC
            LIMIT 50;
        """)

        resumen_anual = [dict(r) for r in db.session.execute(query_resumen, params).mappings().all()]
        resumen_zonas = [dict(r) for r in db.session.execute(query_zonas, params).mappings().all()]
        top_clientes = [dict(r) for r in db.session.execute(query_top_clientes, params).mappings().all()]

        return {
            'success': True,
            'periodo': {'start_year': start_year, 'end_year': end_year},
            'seguridad': {'vista_global': es_global, 'usuario': username},
            'resumen_anual': resumen_anual,
            'resumen_zonas': resumen_zonas,
            'top_clientes': top_clientes
        }

    @staticmethod
    def obtener_detalle_paginado(user_id: int, username: str, user_role: str, start_year: int, end_year: int,
                                  pagina: int = 1, tam_pagina: int = 100, busqueda: str = '') -> dict:
        """
        Detalle consolidado (Año, Mes, Zona, Cliente) con paginación server-side.
        La agregación (GROUP BY) y el filtro de búsqueda corren en PostgreSQL;
        el LIMIT/OFFSET evita que el DTO crezca sin techo a medida que se acumulan años.
        """
        role_upper = str(user_role or '').strip().upper()
        es_global = role_upper in ['ADMIN', 'ADMINISTRACION', 'ADMINISTRADOR', 'GERENCIA']
        es_comercial = not es_global

        pagina = max(1, int(pagina or 1))
        tam_pagina = min(200, max(10, int(tam_pagina or 100)))  # techo duro: nunca mas de 200 filas por respuesta
        offset = (pagina - 1) * tam_pagina

        ventas_expr = ComercialHistoricoService._expr_ajustado_nc('total_ingresos')
        cantidad_expr = ComercialHistoricoService._expr_ajustado_nc('cantidad')
        vendedor_scope = ComercialHistoricoService._resolver_alias_vendedor(username)

        params = {
            'start_year': int(start_year),
            'end_year': int(end_year),
            'es_comercial': es_comercial,
            'vendedor_user': vendedor_scope,
            'vendedor_filtro': '',
            'busqueda': str(busqueda or '').strip(),
            'tam_pagina': tam_pagina,
            'offset': offset
        }

        query = text(f"""
            WITH agrupado AS (
                SELECT
                    EXTRACT(YEAR FROM v.fecha)::INTEGER AS anio,
                    EXTRACT(MONTH FROM v.fecha)::INTEGER AS mes,
                    COALESCE(NULLIF(TRIM(v.zona), ''), 'SIN ZONA') AS zona,
                    COALESCE(NULLIF(TRIM(v.nombres), ''), 'CLIENTE DESCONOCIDO') AS cliente,
                    ROUND(SUM({ventas_expr})::NUMERIC, 2) AS total_ventas,
                    ROUND(SUM({cantidad_expr})::NUMERIC, 2) AS total_unidades,
                    COUNT(v.id)::INTEGER AS total_transacciones
                FROM db_ventas v
                WHERE v.fecha >= MAKE_DATE(:start_year, 1, 1)
                  AND v.fecha <= MAKE_DATE(:end_year, 12, 31)
                  AND {FILTRO_SCOPE_ROL}
                  AND {FILTRO_VENDEDOR_OPCIONAL}
                GROUP BY 1, 2, 3, 4
            )
            SELECT *, COUNT(*) OVER()::INTEGER AS total_registros
            FROM agrupado
            WHERE (:busqueda = '' OR cliente ILIKE '%' || :busqueda || '%' OR zona ILIKE '%' || :busqueda || '%')
            ORDER BY anio ASC, mes ASC, total_ventas DESC
            LIMIT :tam_pagina OFFSET :offset;
        """)

        try:
            filas = [dict(r) for r in db.session.execute(query, params).mappings().all()]
        except Exception as e:
            logger.error(f"[COMERCIAL_SERVICE] Error en detalle paginado: {e}")
            raise e

        total_registros = filas[0]['total_registros'] if filas else 0
        for f in filas:
            f.pop('total_registros', None)

        total_paginas = max(1, (total_registros + tam_pagina - 1) // tam_pagina)

        return {
            'success': True,
            'filas': filas,
            'paginacion': {
                'pagina': pagina,
                'tam_pagina': tam_pagina,
                'total_registros': total_registros,
                'total_paginas': total_paginas
            }
        }

    @staticmethod
    def _formatear_hoja(worksheet, df, columnas_dinero=(), columnas_enteras=(), ancho_col=18):
        """
        Aplica formato profesional a una hoja ya escrita por pandas (index=False):
        encabezado con fondo corporativo, freeze panes, formato contable en columnas
        de dinero/unidades y ancho de columna legible. Reutilizable entre hojas.
        """
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        fondo_header = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        fuente_header = Font(color='FFFFFF', bold=True)
        num_filas = len(df)

        for idx, nombre_col in enumerate(df.columns, start=1):
            letra = get_column_letter(idx)
            celda = worksheet.cell(row=1, column=idx)
            celda.fill = fondo_header
            celda.font = fuente_header
            celda.alignment = Alignment(horizontal='center', vertical='center')
            worksheet.column_dimensions[letra].width = ancho_col

            if nombre_col in columnas_dinero:
                formato = '"$" #,##0'
            elif nombre_col in columnas_enteras:
                formato = '#,##0'
            else:
                formato = None

            if formato:
                for fila in range(2, num_filas + 2):
                    worksheet.cell(row=fila, column=idx).number_format = formato

        worksheet.freeze_panes = 'A2'

    @staticmethod
    def _formatear_hoja_pivot(worksheet, pivot_df, ancho_col=18):
        """
        Variante para hojas escritas con index=True (la pivot Y-o-Y): la primera
        columna es el índice (zona, texto) y el resto son años con valores de dinero.
        """
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        fondo_header = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        fuente_header = Font(color='FFFFFF', bold=True)
        num_filas = len(pivot_df)
        num_columnas = len(pivot_df.columns) + 1  # +1 por la columna de índice (zona)

        for idx in range(1, num_columnas + 1):
            letra = get_column_letter(idx)
            celda = worksheet.cell(row=1, column=idx)
            celda.fill = fondo_header
            celda.font = fuente_header
            celda.alignment = Alignment(horizontal='center', vertical='center')
            worksheet.column_dimensions[letra].width = ancho_col

            if idx > 1:  # columnas de años = dinero; columna 1 = nombre de zona (texto)
                for fila in range(2, num_filas + 2):
                    worksheet.cell(row=fila, column=idx).number_format = '"$" #,##0'

        worksheet.freeze_panes = 'B2'  # columna de zona siempre visible al hacer scroll horizontal

    @staticmethod
    def _query_ytd(user_id: int, username: str, user_role: str, start_year: int, end_year: int,
                    vendedor_filtro: str, corte_dt: date):
        """
        Ejecuta las 4 consultas de agregación YTD (anual, zona, mensual por zona y
        top clientes) 100% en SQL, todas bajo el mismo corte YTD y scope de vendedor
        para que las hojas del Excel sean comparables entre sí.
        El corte YTD compara (mes, dia) por tupla para ser exacto entre años
        bisiestos y no bisiestos (evita el corrimiento que introduce DOY tras el 29-feb).
        """
        role_upper = str(user_role or '').strip().upper()
        es_global = role_upper in ['ADMIN', 'ADMINISTRACION', 'ADMINISTRADOR', 'GERENCIA']
        es_comercial = not es_global

        ventas_expr = ComercialHistoricoService._expr_ajustado_nc('total_ingresos')
        cantidad_expr = ComercialHistoricoService._expr_ajustado_nc('cantidad')
        vendedor_scope = ComercialHistoricoService._resolver_alias_vendedor(username)

        params = {
            'start_year': int(start_year),
            'end_year': int(end_year),
            'es_comercial': es_comercial,
            'vendedor_user': vendedor_scope,
            'vendedor_filtro': str(vendedor_filtro or '').strip(),
            'corte_dt': corte_dt
        }

        where_ytd = f"""
            WHERE EXTRACT(YEAR FROM v.fecha)::INTEGER BETWEEN :start_year AND :end_year
              AND (EXTRACT(MONTH FROM v.fecha), EXTRACT(DAY FROM v.fecha))
                  <= (EXTRACT(MONTH FROM :corte_dt), EXTRACT(DAY FROM :corte_dt))
              AND {FILTRO_SCOPE_ROL}
              AND {FILTRO_VENDEDOR_OPCIONAL}
        """

        query_anual = text(f"""
            SELECT
                EXTRACT(YEAR FROM v.fecha)::INTEGER AS anio,
                ROUND(SUM({ventas_expr})::NUMERIC, 2) AS total_ventas,
                ROUND(SUM({cantidad_expr})::NUMERIC, 2) AS total_unidades,
                COUNT(v.id)::INTEGER AS total_transacciones
            FROM db_ventas v
            {where_ytd}
            GROUP BY EXTRACT(YEAR FROM v.fecha)
            ORDER BY anio ASC;
        """)

        query_zona = text(f"""
            SELECT
                EXTRACT(YEAR FROM v.fecha)::INTEGER AS anio,
                COALESCE(NULLIF(TRIM(v.zona), ''), 'SIN ZONA') AS zona,
                ROUND(SUM({ventas_expr})::NUMERIC, 2) AS total_ventas,
                ROUND(SUM({cantidad_expr})::NUMERIC, 2) AS total_unidades,
                COUNT(v.id)::INTEGER AS total_transacciones
            FROM db_ventas v
            {where_ytd}
            GROUP BY 1, 2
            ORDER BY zona ASC, anio ASC;
        """)

        # Desglose mensual por zona: profundidad temporal dentro de cada año
        query_mensual_zona = text(f"""
            SELECT
                EXTRACT(YEAR FROM v.fecha)::INTEGER AS anio,
                EXTRACT(MONTH FROM v.fecha)::INTEGER AS mes,
                COALESCE(NULLIF(TRIM(v.zona), ''), 'SIN ZONA') AS zona,
                ROUND(SUM({ventas_expr})::NUMERIC, 2) AS total_ventas,
                ROUND(SUM({cantidad_expr})::NUMERIC, 2) AS total_unidades
            FROM db_ventas v
            {where_ytd}
            GROUP BY 1, 2, 3
            ORDER BY anio ASC, mes ASC, zona ASC;
        """)

        # Top clientes del periodo/scope seleccionado (agregado, nunca transacciones sueltas)
        query_top_clientes = text(f"""
            SELECT
                COALESCE(NULLIF(TRIM(v.nombres), ''), 'CLIENTE DESCONOCIDO') AS cliente,
                ROUND(SUM({ventas_expr})::NUMERIC, 2) AS total_ventas,
                ROUND(SUM({cantidad_expr})::NUMERIC, 2) AS total_unidades,
                COUNT(v.id)::INTEGER AS total_transacciones
            FROM db_ventas v
            {where_ytd}
            GROUP BY 1
            ORDER BY total_ventas DESC
            LIMIT 30;
        """)

        try:
            ytd_anual = [dict(r) for r in db.session.execute(query_anual, params).mappings().all()]
            ytd_zona = [dict(r) for r in db.session.execute(query_zona, params).mappings().all()]
            ytd_mensual_zona = [dict(r) for r in db.session.execute(query_mensual_zona, params).mappings().all()]
            top_clientes = [dict(r) for r in db.session.execute(query_top_clientes, params).mappings().all()]
        except Exception as e:
            logger.error(f"[COMERCIAL_SERVICE] Error generando agregación YTD: {e}")
            raise e

        return ytd_anual, ytd_zona, ytd_mensual_zona, top_clientes, es_global

    @staticmethod
    def generar_excel_ytd_stream(user_id: int, username: str, user_role: str, start_year: int, end_year: int,
                                  vendedor_filtro: str = '', fecha_corte=None):
        """
        Genera el .xlsx de analítica comercial YTD/Y-o-Y con formato profesional
        (encabezado corporativo, freeze panes, formato contable) y devuelve
        (buffer, nombre_archivo) listo para pasarle a send_file. Encapsula toda la
        dependencia de pandas/openpyxl para que el controlador quede limpio.
        """
        import pandas as pd
        import io

        corte_dt = ComercialHistoricoService._resolver_fecha_corte(fecha_corte)

        ytd_anual, ytd_zona, ytd_mensual_zona, top_clientes, _ = ComercialHistoricoService._query_ytd(
            user_id=user_id, username=username, user_role=user_role,
            start_year=start_year, end_year=end_year,
            vendedor_filtro=vendedor_filtro, corte_dt=corte_dt
        )

        df_anual = pd.DataFrame(ytd_anual)
        df_zona = pd.DataFrame(ytd_zona)
        df_mensual_zona = pd.DataFrame(ytd_mensual_zona)
        df_top_clientes = pd.DataFrame(top_clientes)

        # Reshape puro (no reagrega nada): pivotea filas ya sumadas por SQL para lectura Y-o-Y.
        if not df_zona.empty:
            pivot_yoy = df_zona.pivot_table(index='zona', columns='anio', values='total_ventas', fill_value=0)
        else:
            pivot_yoy = pd.DataFrame()

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_anual.to_excel(writer, index=False, sheet_name='YTD Anual')
            ComercialHistoricoService._formatear_hoja(
                writer.sheets['YTD Anual'], df_anual,
                columnas_dinero=('total_ventas',), columnas_enteras=('total_unidades', 'total_transacciones')
            )

            df_zona.to_excel(writer, index=False, sheet_name='YTD por Zona')
            ComercialHistoricoService._formatear_hoja(
                writer.sheets['YTD por Zona'], df_zona,
                columnas_dinero=('total_ventas',), columnas_enteras=('total_unidades', 'total_transacciones')
            )

            df_mensual_zona.to_excel(writer, index=False, sheet_name='Desglose Mensual Zona')
            ComercialHistoricoService._formatear_hoja(
                writer.sheets['Desglose Mensual Zona'], df_mensual_zona,
                columnas_dinero=('total_ventas',), columnas_enteras=('total_unidades',)
            )

            df_top_clientes.to_excel(writer, index=False, sheet_name='Top Clientes')
            ComercialHistoricoService._formatear_hoja(
                writer.sheets['Top Clientes'], df_top_clientes,
                columnas_dinero=('total_ventas',), columnas_enteras=('total_unidades', 'total_transacciones'),
                ancho_col=32  # nombres de cliente suelen ser largos
            )

            pivot_yoy.to_excel(writer, sheet_name='YoY Zona (Pivot)')
            if not pivot_yoy.empty:
                ComercialHistoricoService._formatear_hoja_pivot(writer.sheets['YoY Zona (Pivot)'], pivot_yoy)
        output.seek(0)

        vendedor_slug = re.sub(r'[^A-Za-z0-9_-]+', '_', vendedor_filtro.strip())[:40] if vendedor_filtro else 'GLOBAL'
        nombre_archivo = f"Comercial_YTD_{start_year}-{end_year}_{vendedor_slug}_{corte_dt.isoformat()}.xlsx"

        return output, nombre_archivo
