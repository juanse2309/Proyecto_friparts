import logging
from datetime import datetime, date, time

logger = logging.getLogger(__name__)

# Se intenta primero con AM/PM explicito (para resolver sin ambiguedad textos
# legacy tipo '09:00:00 PM'), y luego con 24 horas (formato ya correcto).
_FORMATOS_HORA_TEXTO = [
    '%I:%M:%S %p', '%I:%M %p',
    '%H:%M:%S', '%H:%M',
]


def _parsear_hora_texto(valor_str):
    """
    Intenta interpretar un string de hora en cualquiera de los formatos conocidos
    (12h con AM/PM o 24h) y devuelve un time() sin ambiguedad.
    Si el string es un '%H:%M' de 12 horas SIN indicador AM/PM (ej. formularios
    legacy que guardaron '09:00'), no existe forma de recuperar si era AM o PM;
    se respeta ese valor como 24h (09:00 = 9 de la mañana), unica lectura valida
    sin informacion adicional.
    """
    valor_str = valor_str.strip()
    if not valor_str:
        return None
    for fmt in _FORMATOS_HORA_TEXTO:
        try:
            return datetime.strptime(valor_str, fmt).time()
        except ValueError:
            continue
    return None


def normalizar_hora_24h(valor):
    """Normaliza cualquier valor de hora (datetime, time o string 12h/24h) a 'HH:MM:SS' en formato militar de 24 horas."""
    if valor is None or valor == '':
        return ''
    if isinstance(valor, datetime):
        return valor.strftime('%H:%M:%S')
    if isinstance(valor, time):
        return valor.strftime('%H:%M:%S')

    hora = _parsear_hora_texto(str(valor))
    if hora is None:
        logger.warning(f"[HistorialService] No se pudo normalizar hora a 24h, se deja el valor original: '{valor}'")
        return str(valor).strip()
    return hora.strftime('%H:%M:%S')


def normalizar_fecha_hora_24h(valor):
    """Normaliza un valor de fecha/hora completo a 'YYYY-MM-DD HH:MM:SS' en formato militar de 24 horas."""
    if valor is None or valor == '':
        return ''
    if isinstance(valor, datetime):
        return valor.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(valor, date):
        return datetime.combine(valor, time.min).strftime('%Y-%m-%d %H:%M:%S')

    valor_str = str(valor).strip()
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S', '%d/%m/%Y %H:%M:%S'):
        try:
            return datetime.strptime(valor_str, fmt).strftime('%Y-%m-%d %H:%M:%S')
        except ValueError:
            continue
    return valor_str


def preparar_movimientos_para_excel(movimientos):
    """
    Recibe la lista de movimientos del Historial Global (propiedad exclusiva del
    caller de exportacion, no compartida con nadie mas) y la normaliza in-place
    para exportacion a Excel: los campos de hora quedan estrictamente en formato
    militar de 24 horas (HH:MM:SS), eliminando cualquier ambiguedad AM/PM antes
    de llegar a OpenPyXL.
    Normaliza in-place (no copia la lista completa) para no duplicar el dataset
    en memoria -- con rangos de fecha grandes esa copia fue causa de OOM.
    """
    for mov in movimientos:
        mov['HORA_INICIO'] = normalizar_hora_24h(mov.get('HORA_INICIO'))
        mov['HORA_FIN'] = normalizar_hora_24h(mov.get('HORA_FIN'))
    return movimientos


def generar_excel_historial_global(movimientos):
    """
    Construye el Workbook del Historial Global con 3 hojas:
      - 'Historial Completo': todos los movimientos (comportamiento anterior).
      - 'Inyección': solo Tipo == 'INYECCION'.
      - 'Control PNC': solo Tipo == 'PNC' (agrupa Inyección/Pulido/Ensamble,
        que en el DTO ya comparten el mismo Tipo 'PNC' — ver historial_routes.py).
    Devuelve el buffer BytesIO listo para send_file.

    Usa Workbook(write_only=True): en modo normal OpenPyXL retiene en memoria
    un objeto Cell por cada celda escrita durante toda la vida del Workbook;
    en write_only cada fila se serializa y se libera al hacer ws.append(), lo
    que evita mantener el dataset completo duplicado como arbol de objetos.
    Con rangos de fecha grandes esto fue la causa principal del OOM del server
    (Render free tier, 512MB, ver gunicorn.conf.py).
    """
    from openpyxl import Workbook
    from io import BytesIO

    columnas = [
        'Fecha', 'Hora Inicio', 'Hora Fin', 'Tipo', 'Responsable',
        'Producto', 'Orden Prod.', 'Máquina', 'Cantidad',
        'Peso Bujes (g)', 'Cavidades', 'Duración (s)', 'Tiempo Total (min)', 'Seg/Unidad',
        'Detalle'
    ]
    anchos = [12, 11, 11, 12, 20, 18, 15, 15, 10, 14, 10, 12, 16, 12, 40]

    wb = Workbook(write_only=True)

    ws_completo = wb.create_sheet("Historial Completo")
    _escribir_hoja_historial(ws_completo, movimientos, columnas, anchos)

    movimientos_inyeccion = [m for m in movimientos if m.get('Tipo') == 'INYECCION']
    ws_iny = wb.create_sheet("Inyección")
    _escribir_hoja_historial(ws_iny, movimientos_inyeccion, columnas, anchos)

    movimientos_pnc = [m for m in movimientos if m.get('Tipo') == 'PNC']
    ws_pnc = wb.create_sheet("Control PNC")
    _escribir_hoja_historial(ws_pnc, movimientos_pnc, columnas, anchos)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


_COLUMNAS_TEXTO_IZQ = {4, 5, 6, 7, 8, 15}  # Tipo, Responsable, Producto, Orden, Máquina, Detalle


def _escribir_hoja_historial(ws, movimientos, columnas, anchos):
    """
    Escribe cabecera en negrita, filas saneadas, cebreado y anchos en una hoja
    del Historial Global (worksheet en modo write_only: las celdas se arman
    fila por fila con WriteOnlyCell y se entregan via ws.append()).
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.utils import get_column_letter

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D5D8DC'),
        right=Side(style='thin', color='D5D8DC'),
        top=Side(style='thin', color='D5D8DC'),
        bottom=Side(style='thin', color='D5D8DC')
    )
    zebra_fill = PatternFill(start_color='F2F3F4', end_color='F2F3F4', fill_type='solid')
    data_align = Alignment(horizontal='center', vertical='center')
    text_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # En modo write_only el anchor/freeze deben fijarse ANTES del primer
    # ws.append(): el writer streamea <cols>/panes al abrir la hoja, y una
    # vez que arrancó <sheetData> ya no puede insertarlos (quedan en 13.0
    # default / sin freeze, sin error visible).
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    fila_header = []
    for titulo_col in columnas:
        cell = WriteOnlyCell(ws, value=titulo_col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        fila_header.append(cell)
    ws.append(fila_header)

    for row_idx, r in enumerate(movimientos, 2):
        fila = [
            r.get('Fecha', ''),
            r.get('HORA_INICIO', ''),
            r.get('HORA_FIN', ''),
            r.get('Tipo', ''),
            r.get('Responsable', ''),
            r.get('Producto', ''),
            r.get('Orden', ''),
            r.get('maquina', 'N/A'),
            r.get('Cant', 0),
            r.get('peso_bujes'),
            r.get('cavidades'),
            r.get('duracion_segundos'),
            r.get('tiempo_total_minutos'),
            r.get('segundos_por_unidad'),
            r.get('Detalle', '')
        ]

        es_par = (row_idx % 2 == 0)
        fila_celdas = []
        for col_idx, valor in enumerate(fila, 1):
            # Purgar estrictamente cualquier representación de nulo a None para celda vacía en Excel
            if valor is None or (isinstance(valor, float) and (valor != valor)) or str(valor).strip().lower() in ('nan', 'none', 'null'):
                cell_val = None
            else:
                cell_val = valor

            cell = WriteOnlyCell(ws, value=cell_val)
            cell.border = thin_border

            # Columnas 2 y 3 = Hora Inicio / Hora Fin: forzar formato Texto
            # para que OpenPyXL/Excel nunca reinterprete el string 24h
            # normalizado como una hora 12h dependiente del locale.
            if col_idx in (2, 3):
                cell.number_format = '@'

            cell.alignment = text_align if col_idx in _COLUMNAS_TEXTO_IZQ else data_align

            if es_par:
                cell.fill = zebra_fill

            fila_celdas.append(cell)

        ws.append(fila_celdas)
